# -*- coding: utf-8 -*-
"""人情礼金系统 Blueprint"""
import os, csv, io, sqlite3, threading, time as time_mod
from flask import Blueprint, request, jsonify, Response

from .utils import make_logger, make_db

# openpyxl 按需懒加载，避免启动时阻塞
_openpyxl = None
def _get_openpyxl():
    global _openpyxl
    if _openpyxl is None:
        import openpyxl
        _openpyxl = openpyxl
    return _openpyxl

bp = Blueprint('renqing', __name__, url_prefix='/api/renqing')

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RENQING_DB_PATH = os.path.join(BASE_DIR, '人情', 'gifts.db')
LOG_FILE = os.path.join(BASE_DIR, '人情', 'renqing.log')

_log = make_logger(LOG_FILE)
SCHEMA_VERSION = 4


def init_db():
    try:
        os.makedirs(os.path.dirname(RENQING_DB_PATH), exist_ok=True)
        conn = _get_db()
        conn.row_factory = sqlite3.Row
        conn.executescript('''
            CREATE TABLE IF NOT EXISTS events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                event_type TEXT DEFAULT '日常',
                event_date TEXT,
                locked INTEGER DEFAULT 0,
                page_size INTEGER DEFAULT 12
            );
            CREATE TABLE IF NOT EXISTS records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                event_id INTEGER NOT NULL,
                date TEXT,
                name TEXT NOT NULL,
                amount REAL NOT NULL,
                note TEXT,
                batch INTEGER DEFAULT 1,
                direction TEXT DEFAULT '收',
                FOREIGN KEY (event_id) REFERENCES events(id)
            );
            CREATE INDEX IF NOT EXISTS idx_records_event_id ON records(event_id);
            CREATE INDEX IF NOT EXISTS idx_records_name ON records(name);
            CREATE INDEX IF NOT EXISTS idx_records_batch ON records(batch);
            CREATE TABLE IF NOT EXISTS schema_version (
                version INTEGER PRIMARY KEY
            );
        ''')
        ver = conn.execute('SELECT version FROM schema_version LIMIT 1').fetchone()
        cur_ver = ver['version'] if ver else 0

        if cur_ver < SCHEMA_VERSION:
            try: conn.execute('ALTER TABLE events ADD COLUMN locked INTEGER DEFAULT 0')
            except sqlite3.OperationalError: pass
            try: conn.execute("ALTER TABLE records ADD COLUMN direction TEXT DEFAULT '收'")
            except sqlite3.OperationalError: pass
            try: conn.execute("ALTER TABLE events ADD COLUMN event_type TEXT DEFAULT '事件'")
            except sqlite3.OperationalError: pass
            try: conn.execute("ALTER TABLE events ADD COLUMN page_size INTEGER DEFAULT 12")
            except sqlite3.OperationalError: pass
            try: conn.execute("ALTER TABLE events ADD COLUMN archived INTEGER DEFAULT 0")
            except sqlite3.OperationalError: pass
            conn.execute("UPDATE records SET direction='送' WHERE direction IS NULL OR direction=''")
            conn.execute("UPDATE records SET direction='送' WHERE event_id IN (SELECT id FROM events WHERE name='流水')")
            conn.execute("UPDATE records SET batch=CAST(substr(date,1,4) AS INTEGER) WHERE event_id IN (SELECT id FROM events WHERE name='流水') AND date IS NOT NULL AND date!='' AND CAST(substr(date,1,4) AS INTEGER)>0 AND batch!=CAST(substr(date,1,4) AS INTEGER)")
            # 将已有"流水"事件标记为"日常"类型
            conn.execute("UPDATE events SET event_type='日常' WHERE name='流水' AND (event_type='' OR event_type='事件' OR event_type IS NULL)")
            conn.execute("UPDATE records SET direction='送' WHERE event_id IN (SELECT id FROM events WHERE event_type='日常') AND direction!='送'")
            # 修复：名为"受伤"的事件应是收礼事件，纳入事件管理页
            conn.execute("UPDATE events SET event_type='事件' WHERE name='受伤' AND event_type='日常'")
            conn.execute("UPDATE records SET direction='收' WHERE event_id IN (SELECT id FROM events WHERE name='受伤') AND direction!='收'")
            conn.execute('DELETE FROM schema_version')
            conn.execute('INSERT INTO schema_version VALUES (?)', (SCHEMA_VERSION,))
            conn.commit()
            # Excel 历史数据导入移到后台线程，避免阻塞启动（仅首次 records 为空时执行）
            xlsx_path = os.path.join(BASE_DIR, '人情', '人情记录.xlsx')
            if os.path.exists(xlsx_path) and conn.execute('SELECT COUNT(*) FROM records').fetchone()[0] == 0:
                conn.close()
                threading.Thread(target=_import_excel_async, args=(xlsx_path,), daemon=True).start()
                return
        conn.commit()
        conn.close()
    except Exception as e:
        import traceback
        print(f'[init_renqing_db ERROR] {e}\n{traceback.format_exc()}', flush=True)


def _import_excel_async(xlsx_path):
    """后台线程：从 Excel 导入历史数据（首次运行且 records 为空时）"""
    try:
        time_mod.sleep(5)  # 等待 Flask 完全启动后再开始导入
        wb = _get_openpyxl().load_workbook(xlsx_path)
        conn = _get_db()
        for sname in wb.sheetnames:
            ws = wb[sname]
            conn.execute("INSERT OR IGNORE INTO events (name, event_type) VALUES (?,?)",
                         (sname, '日常' if sname == '流水' else '事件'))
            eid = conn.execute("SELECT id FROM events WHERE name=?", (sname,)).fetchone()[0]
            fmt_date = lambda v: v.strftime('%Y-%m-%d') if hasattr(v, 'strftime') else str(v) if v else ''
            for row in ws.iter_rows(min_row=2, values_only=True):
                vals = row + (None,)*4
                dir_flag = '送' if sname == '流水' else '收'
                if sname == '流水':
                    d, n, a, note = vals[:4]
                    if n and a:
                        conn.execute('INSERT INTO records (event_id,date,name,amount,note,direction) VALUES (?,?,?,?,?,?)',
                                     (eid, fmt_date(d), str(n), float(a), str(note or ''), dir_flag))
                elif sname == '爷爷仙逝':
                    n, a, note, b = vals[:4]
                    if n and a:
                        conn.execute('INSERT INTO records (event_id,name,amount,note,batch,direction) VALUES (?,?,?,?,?,?)',
                                     (eid, str(n), float(a), str(note or ''), int(b or 1), dir_flag))
                else:
                    t, a, b = vals[:3]
                    if t and a:
                        conn.execute('INSERT INTO records (event_id,name,amount,batch,direction) VALUES (?,?,?,?,?)',
                                     (eid, str(t), float(a), int(b or 1), dir_flag))
        conn.commit()
        conn.close()
        print('[renqing] Excel 历史数据导入完成', flush=True)
    except Exception as e:
        import traceback
        print(f'[_import_excel_async ERROR] {e}\n{traceback.format_exc()}', flush=True)


_get_db = make_db(RENQING_DB_PATH)


def _lock_check(c, eid):
    r = c.execute("SELECT locked FROM events WHERE id=?", (eid,)).fetchone()
    if r and r['locked']:
        return jsonify({'success': False, 'error': '事件已锁定'}), 403
    return None


def _adjust_for_flow(c, eid, d):
    evt = c.execute('SELECT name, event_type FROM events WHERE id=?', (eid,)).fetchone()
    if evt and (evt['name'] == '流水' or evt['event_type'] == '日常'):
        d['direction'] = '送'
        if d.get('date'): d['batch'] = int(str(d['date'])[:4])
    return d


def _auto_batch(c, eid):
    """事件收礼：自动分配页，根据事件配置的page_size"""
    evt = c.execute('SELECT page_size FROM events WHERE id=?', (eid,)).fetchone()
    ps = evt['page_size'] if evt and evt['page_size'] else 12
    max_batch = c.execute('SELECT MAX(batch) as mb FROM records WHERE event_id=?', (eid,)).fetchone()['mb'] or 0
    if max_batch == 0:
        return 1
    cnt = c.execute('SELECT COUNT(*) as c FROM records WHERE event_id=? AND batch=?', (eid, max_batch)).fetchone()['c']
    return max_batch + 1 if cnt >= ps else max_batch


def _is_flow_event(c, eid):
    evt = c.execute('SELECT name, event_type FROM events WHERE id=?', (eid,)).fetchone()
    if not evt: return False
    return evt['name'] == '流水' or evt['event_type'] == '日常'


# ==================== 仪表盘API ====================

@bp.route('/dashboard')
def dashboard():
    c = _get_db()
    try:
        ov = dict(c.execute('''SELECT COUNT(*) as total_records, COUNT(DISTINCT name) as person_cnt,
            COALESCE(SUM(amount),0) as total_amount,
            COALESCE(SUM(CASE WHEN direction='收' THEN amount ELSE 0 END),0) as total_rec,
            COALESCE(SUM(CASE WHEN direction='送' THEN amount ELSE 0 END),0) as total_send
            FROM records''').fetchone())
        recent = [dict(r) for r in c.execute('''SELECT r.*, e.name as event_name
            FROM records r JOIN events e ON r.event_id=e.id
            ORDER BY r.id DESC LIMIT 10''').fetchall()]
        events_summary = [dict(r) for r in c.execute('''SELECT e.*,
            COUNT(r.id) as record_count,
            COALESCE(SUM(r.amount),0) as total_amount
            FROM events e LEFT JOIN records r ON e.id=r.event_id
            GROUP BY e.id ORDER BY e.id''').fetchall()]
        balance = [dict(r) for r in c.execute('''SELECT r.name,
            COALESCE(SUM(CASE WHEN r.direction='收' THEN r.amount ELSE 0 END),0) as total_rec,
            COALESCE(SUM(CASE WHEN r.direction='送' THEN r.amount ELSE 0 END),0) as total_send
            FROM records r GROUP BY r.name
            HAVING ABS(COALESCE(SUM(CASE WHEN r.direction='收' THEN r.amount ELSE 0 END),0) -
                       COALESCE(SUM(CASE WHEN r.direction='送' THEN r.amount ELSE 0 END),0)) > 0
            ORDER BY ABS(COALESCE(SUM(CASE WHEN r.direction='收' THEN r.amount ELSE 0 END),0) -
                         COALESCE(SUM(CASE WHEN r.direction='送' THEN r.amount ELSE 0 END),0)) DESC
            LIMIT 10''').fetchall()]
        return jsonify({'overview': ov, 'recent': recent,
                        'events': events_summary, 'balance': balance})
    finally:
        c.close()


# ==================== 事件API ====================

@bp.route('/events', methods=['GET'])
def list_events():
    c = _get_db()
    try:
        rows = c.execute('''SELECT e.*,COUNT(r.id) as record_count,COALESCE(SUM(r.amount),0) as total_amount
            FROM events e LEFT JOIN records r ON e.id=r.event_id GROUP BY e.id ORDER BY CASE WHEN e.event_type='日常' THEN 0 ELSE 1 END, e.event_date DESC, e.id DESC''').fetchall()
        events = [dict(r) for r in rows]
        # 批次统计
        eids = [e['id'] for e in events]
        if eids:
            ph = ','.join('?' for _ in eids)
            br = c.execute(f'''SELECT event_id, batch, COUNT(*) as cnt, SUM(amount) as total
                FROM records WHERE event_id IN ({ph}) AND batch IS NOT NULL
                GROUP BY event_id, batch ORDER BY event_id, batch''', eids).fetchall()
            bm = {}
            for b in br:
                bm.setdefault(b['event_id'], []).append({'batch': b['batch'], 'cnt': b['cnt'], 'total': b['total']})
            for e in events:
                e['batch_stats'] = bm.get(e['id'], [])
        else:
            for e in events:
                e['batch_stats'] = []
        return jsonify(events)
    finally:
        c.close()


@bp.route('/events', methods=['POST'])
def add_event():
    d = request.get_json(silent=True) or {}
    name = (d.get('name') or '').strip()
    if not name:
        return jsonify({'success': False, 'error': '事件名称不能为空'}), 400

    c = _get_db()
    try:
        c.execute("INSERT INTO events (name,event_type,event_date,page_size) VALUES (?,?,?,?)",
                  (name, d.get('event_type','事件'), d.get('event_date',''), d.get('page_size', 12)))
        c.commit()
        _log(f'新增事件: {name}')
        return jsonify({'success': True})
    except sqlite3.IntegrityError:
        return jsonify({'success': False, 'error': f'事件"{name}"已存在，请使用其他名称'}), 400
    except Exception as e:
        _log(f'新增事件失败: {name} - {e}')
        return jsonify({'success': False, 'error': f'保存失败: {str(e)[:100]}'}), 500
    finally:
        try: c.close()
        except: pass


@bp.route('/events/<int:eid>', methods=['PUT'])
def update_event(eid):
    d = request.get_json(silent=True) or {}
    c = _get_db()
    try:
        err = _lock_check(c, eid)
        if err: return err
        c.execute("UPDATE events SET name=?,event_type=?,event_date=?,page_size=? WHERE id=?",
                  (d.get('name',''), d.get('event_type','事件'), d.get('event_date',''), d.get('page_size', 12), eid))
        c.commit()
        _log(f'更新事件: {d["name"]}')
        return jsonify({'success': True})
    except sqlite3.IntegrityError:
        return jsonify({'success': False, 'error': f'事件名称"{d.get("name","")}"已存在，请使用其他名称'}), 400
    except Exception as e:
        _log(f'更新事件失败: {d.get("name","")} - {e}')
        return jsonify({'success': False, 'error': f'保存失败: {str(e)[:100]}'}), 500
    finally:
        try: c.close()
        except: pass


@bp.route('/events/<int:eid>/lock', methods=['POST'])
def toggle_lock(eid):
    c = _get_db()
    try:
        r = c.execute("SELECT locked FROM events WHERE id=?", (eid,)).fetchone()
        if not r: return jsonify({'success': False, 'error': '不存在'}), 404
        new_lock = 0 if r['locked'] else 1
        c.execute("UPDATE events SET locked=? WHERE id=?", (new_lock, eid))
        c.commit()
        _log(f'{"锁定" if new_lock else "解锁"}事件: ID:{eid}')
        return jsonify({'success': True})
    finally:
        c.close()


@bp.route('/events/<int:eid>/archive', methods=['POST'])
def toggle_archive(eid):
    c = _get_db()
    try:
        r = c.execute("SELECT archived FROM events WHERE id=?", (eid,)).fetchone()
        if not r: return jsonify({'success': False, 'error': '不存在'}), 404
        new_archive = 0 if r['archived'] else 1
        c.execute("UPDATE events SET archived=? WHERE id=?", (new_archive, eid))
        c.commit()
        _log(f'{"归档" if new_archive else "取消归档"}事件: ID:{eid}')
        return jsonify({'success': True, 'archived': bool(new_archive)})
    finally:
        c.close()


@bp.route('/events/<int:eid>', methods=['DELETE'])
def delete_event(eid):
    c = _get_db()
    try:
        err = _lock_check(c, eid)
        if err: return err
        evt = c.execute("SELECT name FROM events WHERE id=?", (eid,)).fetchone()
        evt_name = evt['name'] if evt else str(eid)
        c.execute('DELETE FROM records WHERE event_id=?', (eid,))
        c.execute('DELETE FROM events WHERE id=?', (eid,))
        c.commit()
        _log(f'删除事件: {evt_name}')
        return jsonify({'success': True})
    finally:
        c.close()


# ==================== 记录API ====================

@bp.route('/records', methods=['GET'])
def list_records():
    c = _get_db(); p = []
    try:
        q = '''SELECT r.*,e.name as event_name,e.event_date,
               COALESCE(r.date,e.event_date) as display_date
               FROM records r JOIN events e ON r.event_id=e.id WHERE 1=1'''
        eid = request.args.get('event_id')
        if eid: q += ' AND r.event_id=?'; p.append(eid)
        s = request.args.get('search')
        if s: q += ' AND (r.name LIKE ? OR r.note LIKE ?)'; p += [f'%{s}%']*2
        evt_name = None
        if eid:
            evt = c.execute('SELECT name, event_type FROM events WHERE id=?', (eid,)).fetchone()
            if evt: evt_name = evt['name']
        if evt_name and (evt_name == '流水' or evt['event_type'] == '日常'):
            q += ' ORDER BY r.batch DESC, r.date DESC NULLS LAST, r.id DESC'
        else:
            q += ' ORDER BY r.batch ASC, r.id ASC'
        rows = c.execute(q, p).fetchall()
        return jsonify([dict(r) for r in rows])
    finally:
        c.close()


@bp.route('/records', methods=['POST'])
def add_record():
    d = request.get_json(silent=True) or {}; c = _get_db()
    try:
        eid = d.get('event_id')
        if not eid:
            return jsonify({'success': False, 'error': '缺少事件ID'}), 400
        name = (d.get('name') or '').strip()
        if not name:
            return jsonify({'success': False, 'error': '姓名不能为空'}), 400
        amount = d.get('amount')
        if amount is None or amount <= 0:
            return jsonify({'success': False, 'error': '金额必须大于0'}), 400
        # 事件收礼(非日常)：阻止向已锁定事件新增记录
        evt_info = c.execute('SELECT event_type,locked FROM events WHERE id=?', (eid,)).fetchone()
        if not evt_info:
            return jsonify({'success': False, 'error': '事件不存在'}), 404
        if evt_info['event_type'] != '日常' and evt_info['locked']:
            return jsonify({'success': False, 'error': '事件已锁定，无法新增记录'}), 403
        d = _adjust_for_flow(c, eid, d)
        # 事件收礼：若前端显式指定batch则使用；否则自动分配（保证每页独立录入）
        if not _is_flow_event(c, eid):
            if d.get('batch') is None:
                d['batch'] = _auto_batch(c, eid)
        date_val = d.get('date','')
        if not date_val:
            evt = c.execute('SELECT event_date FROM events WHERE id=?', (eid,)).fetchone()
            if evt and evt['event_date']:
                date_val = evt['event_date']
        c.execute('INSERT INTO records (event_id,date,name,amount,note,batch,direction) VALUES (?,?,?,?,?,?,?)',
                  (eid, date_val, name, amount, d.get('note',''), d.get('batch',1), d.get('direction','收')))
        c.commit()
        _log(f'新增记录: {name} {amount}元')
        return jsonify({'success': True})
    finally:
        c.close()


@bp.route('/records/<int:rid>', methods=['PUT'])
def update_record(rid):
    d = request.get_json(silent=True) or {}; c = _get_db()
    try:
        r = c.execute("SELECT event_id FROM records WHERE id=?", (rid,)).fetchone()
        if not r:
            return jsonify({'success': False, 'error': '记录不存在'}), 404
        err = _lock_check(c, r['event_id'])
        if err: return err
        # 获取目标事件（优先用请求中的event_id，否则保持原记录的事件）
        target_eid = d.get('event_id', r['event_id'])
        d = _adjust_for_flow(c, target_eid, d)
        date_val = d.get('date','')
        if not date_val:
            evt = c.execute('SELECT event_date FROM events WHERE id=?', (target_eid,)).fetchone()
            if evt and evt['event_date']:
                date_val = evt['event_date']
        amount = d.get('amount', 0)
        try:
            amount = float(amount)
            if amount <= 0:
                return jsonify({'success': False, 'error': '金额必须大于0'}), 400
        except (ValueError, TypeError):
            return jsonify({'success': False, 'error': '金额格式无效'}), 400
        c.execute('UPDATE records SET event_id=?,date=?,name=?,amount=?,note=?,batch=?,direction=? WHERE id=?',
                  (target_eid, date_val, d.get('name',''), amount, d.get('note',''), d.get('batch',1), d.get('direction','收'), rid))
        c.commit()
        _log(f'更新记录: {d["name"]} (ID:{rid})')
        return jsonify({'success': True})
    finally:
        c.close()


@bp.route('/records/<int:rid>', methods=['DELETE'])
def delete_record(rid):
    c = _get_db()
    try:
        r = c.execute("SELECT event_id FROM records WHERE id=?", (rid,)).fetchone()
        if r:
            err = _lock_check(c, r['event_id'])
            if err: return err
        c.execute('DELETE FROM records WHERE id=?', (rid,))
        c.commit()
        _log(f'删除记录: ID:{rid}')
        return jsonify({'success': True})
    finally:
        c.close()


# ==================== 统计API ====================

@bp.route('/stats')
def stats():
    c = _get_db()
    try:
        eid = request.args.get('event_id')
        if eid:
            try:
                eid = int(eid)
            except (ValueError, TypeError):
                return jsonify({'success': False, 'error': '无效的事件ID'}), 400
            w = 'WHERE r.event_id=?'
            p = [eid]
        else:
            w = ''
            p = []
        row = c.execute(f'''SELECT COALESCE(SUM(r.amount),0) as total,
            COALESCE(SUM(CASE WHEN r.direction='收' THEN r.amount ELSE 0 END),0) as total_rec,
            COALESCE(SUM(CASE WHEN r.direction='送' THEN r.amount ELSE 0 END),0) as total_send,
            COUNT(*) as cnt FROM records r {w}''', p).fetchone()
        ns = c.execute(f'SELECT r.name,SUM(r.amount) as total,COUNT(*) as cnt FROM records r {w} GROUP BY r.name ORDER BY total DESC', p).fetchall()
        return jsonify({'total': row['total'], 'total_rec': row['total_rec'], 'total_send': row['total_send'],
                        'count': row['cnt'], 'name_stats': [dict(n) for n in ns]})
    finally:
        c.close()


@bp.route('/suggestions')
def suggestions():
    q = request.args.get('q', '').strip()
    if not q: return jsonify([])
    c = None
    try:
        c = _get_db()
        rows = c.execute('SELECT DISTINCT name FROM records WHERE name LIKE ? ORDER BY name LIMIT 20', (f'%{q}%',)).fetchall()
        return jsonify([r['name'] for r in rows])
    except Exception as e:
        return jsonify([])  # 建议功能失败不阻塞用户操作，静默返回空
    finally:
        if c:
            c.close()


@bp.route('/person')
def search_person():
    c = None
    try:
        name = request.args.get('name', '').strip()
        if not name: return jsonify({'success': False, 'error': '请输入姓名'}), 400
        c = _get_db()
        rows = c.execute('''SELECT r.*,e.name as event_name,e.locked as event_locked,e.event_date,e.event_type,
            COALESCE(r.date,e.event_date) as display_date
            FROM records r JOIN events e ON r.event_id=e.id
            WHERE r.name LIKE ? ORDER BY e.id, r.date DESC NULLS LAST''',
            (f'%{name}%',)).fetchall()
        evts = {}
        for r in rows:
            eid = r['event_id']
            if eid not in evts:
                evts[eid] = {'event_name': r['event_name'], 'event_id': eid,
                             'event_date': r['event_date'], 'event_locked': r['event_locked'],
                             'event_type': r['event_type'], 'records': [], 'total': 0}
            evts[eid]['records'].append(dict(r))
            evts[eid]['total'] += r['amount']
        return jsonify({'name': name, 'total_amount': sum(r['amount'] for r in rows),
                        'total_rec': sum(r['amount'] for r in rows if r['direction'] == '收'),
                        'total_send': sum(r['amount'] for r in rows if r['direction'] == '送'),
                        'total_count': len(rows), 'events': list(evts.values())})
    except Exception as e:
        return jsonify({'success': False, 'error': f'查询失败: {str(e)}'}), 500
    finally:
        if c: c.close()


@bp.route('/merge-names', methods=['POST'])
def merge_names():
    c = None
    try:
        source = (request.form.get('source') or '').strip()
        target = (request.form.get('target') or '').strip()
        if not source or not target:
            return jsonify({'success': False, 'error': '参数不完整'}), 400
        if source == target:
            return jsonify({'success': False, 'error': '两个名字相同'}), 400
        c = _get_db()
        # 检查是否有锁定事件中的记录
        locked = c.execute('''SELECT e.name FROM records r
            JOIN events e ON r.event_id=e.id
            WHERE r.name=? AND e.locked=1 LIMIT 1''', (source,)).fetchone()
        if locked:
            return jsonify({'success': False, 'error': f'"{source}"有记录属于已锁定事件"{locked["name"]}"，不能合并'}), 400
        cur = c.execute('UPDATE records SET name=? WHERE name=?', (target, source))
        cnt = cur.rowcount; c.commit()
        _log(f'合并姓名: {source} -> {target} ({cnt}条)')
        return jsonify({'success': True, 'merged': cnt})
    except Exception as e:
        return jsonify({'success': False, 'error': f'合并失败: {str(e)}'}), 500
    finally:
        if c: c.close()


# ==================== 分析API ====================

@bp.route('/analysis')
def analysis():
    c = _get_db()
    try:
        by_person = [dict(r) for r in c.execute('''SELECT r.name,SUM(r.amount) as total,COUNT(*) as cnt,
            SUM(CASE WHEN r.direction='收' THEN r.amount ELSE 0 END) as total_rec,
            SUM(CASE WHEN r.direction='送' THEN r.amount ELSE 0 END) as total_send
            FROM records r GROUP BY r.name ORDER BY total DESC''').fetchall()]
        by_event = [dict(r) for r in c.execute('''SELECT e.name,COUNT(r.id) as cnt,COALESCE(SUM(r.amount),0) as total,
            COALESCE(SUM(CASE WHEN r.direction='收' THEN r.amount ELSE 0 END),0) as total_rec,
            COALESCE(SUM(CASE WHEN r.direction='送' THEN r.amount ELSE 0 END),0) as total_send
            FROM events e LEFT JOIN records r ON e.id=r.event_id GROUP BY e.id ORDER BY total DESC''').fetchall()]
        by_month = [dict(r) for r in c.execute('''SELECT substr(r.date,1,7) as month,SUM(r.amount) as total,COUNT(*) as cnt,
            SUM(CASE WHEN r.direction='收' THEN r.amount ELSE 0 END) as total_rec,
            SUM(CASE WHEN r.direction='送' THEN r.amount ELSE 0 END) as total_send
            FROM records r JOIN events e ON r.event_id=e.id WHERE (e.event_type='日常' OR e.name='流水') AND r.date!='' GROUP BY month ORDER BY month DESC''').fetchall()]
        by_year = [dict(r) for r in c.execute('''SELECT substr(r.date,1,4) as year,SUM(r.amount) as total,COUNT(*) as cnt,
            SUM(CASE WHEN r.direction='收' THEN r.amount ELSE 0 END) as total_rec,
            SUM(CASE WHEN r.direction='送' THEN r.amount ELSE 0 END) as total_send
            FROM records r JOIN events e ON r.event_id=e.id WHERE e.name='流水' AND r.date!='' GROUP BY year ORDER BY year''').fetchall()]
        dist_rows = c.execute('''SELECT
            CASE WHEN amount<100 THEN '0-100' WHEN amount<200 THEN '100-200' WHEN amount<300 THEN '200-300'
            WHEN amount<500 THEN '300-500' WHEN amount<1000 THEN '500-1000' ELSE '1000+' END as range_label,
            COUNT(*) as count FROM records GROUP BY range_label ORDER BY range_label''').fetchall()
        dist = [{'range': r['range_label'], 'count': r['count']} for r in dist_rows]
        batch = [dict(r) for r in c.execute('''SELECT e.name as event_name,e.id as event_id,r.batch,
            COUNT(*) as cnt,SUM(r.amount) as total FROM records r JOIN events e ON r.event_id=e.id
            GROUP BY e.id,r.batch ORDER BY e.id,r.batch''').fetchall()]
        raw = c.execute('''SELECT r.name,r.event_id,e.name as event_name,r.amount,r.date,r.note,r.batch,r.direction
            FROM records r JOIN events e ON r.event_id=e.id ORDER BY r.name,r.event_id''').fetchall()
        pm = {}
        for row in raw:
            n, eid = row['name'], row['event_id']
            if n not in pm: pm[n] = {'name': n, 'events': {}}
            if eid not in pm[n]['events']:
                pm[n]['events'][eid] = {'event_name': row['event_name'], 'records': [], 'total': 0}
            neg_f = row['direction'] == '送'
            pm[n]['events'][eid]['records'].append({
                'amount': row['amount'], 'date': row['date'], 'note': row['note'],
                'batch': row['batch'], 'direction': row['direction'] or '收', 'is_negative': neg_f
            })
            pm[n]['events'][eid]['total'] += (-row['amount'] if neg_f else row['amount'])
        cross, single = [], []
        for v in pm.values():
            evts = v['events']
            if len(evts) > 1:
                net = sum(e['total'] for e in evts.values())
                cross.append({'name': v['name'], 'event_cnt': len(evts), 'events': list(evts.values()), 'net_total': net})
            elif len(evts) == 1:
                eid = list(evts.keys())[0]
                e = evts[eid]
                single.append({'name': v['name'], 'event_name': e['event_name'], 'total': e['total'], 'count': len(e['records'])})
        cross.sort(key=lambda x: abs(x['net_total']), reverse=True)
        single.sort(key=lambda x: x['total'], reverse=True)
        ov = dict(c.execute('''SELECT COUNT(DISTINCT event_id) as event_cnt,COUNT(DISTINCT name) as person_cnt,
            SUM(amount) as total,COUNT(*) as record_cnt,
            SUM(CASE WHEN direction='收' THEN amount ELSE 0 END) as total_rec,
            SUM(CASE WHEN direction='送' THEN amount ELSE 0 END) as total_send FROM records''').fetchone())
        return jsonify({'overview': ov, 'by_person': by_person, 'by_event': by_event, 'by_month': by_month,
                        'by_year': by_year, 'distribution': dist, 'batch_summary': batch,
                        'cross_person': cross, 'single_person': single})
    finally:
        c.close()


@bp.route('/compare')
def compare_events():
    eid1 = request.args.get('event_id1')
    eid2 = request.args.get('event_id2')
    if not eid1 or not eid2:
        return jsonify({'success': False, 'error': '请选择两个事件'}), 400
    c = _get_db()
    try:
        e1 = dict(c.execute('SELECT * FROM events WHERE id=?', (eid1,)).fetchone() or {})
        e2 = dict(c.execute('SELECT * FROM events WHERE id=?', (eid2,)).fetchone() or {})
        if not e1 or not e2:
            return jsonify({'success': False, 'error': '事件不存在'}), 404
        r1_rows = c.execute('SELECT name,amount,direction FROM records WHERE event_id=?', (eid1,)).fetchall()
        r2_rows = c.execute('SELECT name,amount,direction FROM records WHERE event_id=?', (eid2,)).fetchall()
        r1, r2 = {}, {}
        for r in r1_rows:
            if r['name'] not in r1: r1[r['name']] = {'amount':0,'direction':r['direction']}
            r1[r['name']]['amount'] += r['amount']
        for r in r2_rows:
            if r['name'] not in r2: r2[r['name']] = {'amount':0,'direction':r['direction']}
            r2[r['name']]['amount'] += r['amount']
        n1, n2 = set(r1.keys()), set(r2.keys())
        common = sorted([{'name': n, 'amount1': r1[n]['amount'], 'amount2': r2[n]['amount'],
                          'diff': r2[n]['amount'] - r1[n]['amount'],
                          'dir1': r1[n]['direction'], 'dir2': r2[n]['direction']} for n in (n1 & n2)], key=lambda x: x['name'])
        only1 = sorted([{'name': n, 'amount': r1[n]['amount'], 'direction': r1[n]['direction']} for n in (n1 - n2)], key=lambda x: x['name'])
        only2 = sorted([{'name': n, 'amount': r2[n]['amount'], 'direction': r2[n]['direction']} for n in (n2 - n1)], key=lambda x: x['name'])
        return jsonify({
            'event1': {'id': e1['id'], 'name': e1['name'], 'count': len(r1_rows), 'total': sum(v['amount'] for v in r1.values())},
            'event2': {'id': e2['id'], 'name': e2['name'], 'count': len(r2_rows), 'total': sum(v['amount'] for v in r2.values())},
            'common': common, 'only_in_1': only1, 'only_in_2': only2,
            'summary': {'common': len(common), 'only1': len(only1), 'only2': len(only2)}
        })
    finally:
        c.close()


@bp.route('/balance')
def balance():
    c = _get_db()
    try:
        rows = c.execute('''SELECT r.name,
            SUM(CASE WHEN r.direction='收' THEN r.amount ELSE 0 END) as total_rec,
            SUM(CASE WHEN r.direction='送' THEN r.amount ELSE 0 END) as total_send,
            COUNT(*) as cnt
            FROM records r JOIN events e ON r.event_id=e.id
            GROUP BY r.name ORDER BY ABS(SUM(CASE WHEN r.direction='收' THEN r.amount ELSE 0 END) - SUM(CASE WHEN r.direction='送' THEN r.amount ELSE 0 END)) DESC''').fetchall()
        result = []
        for r in rows:
            net = (r['total_rec'] or 0) - (r['total_send'] or 0)
            result.append({'name': r['name'], 'rec': r['total_rec'] or 0, 'send': r['total_send'] or 0,
                           'net': net, 'cnt': r['cnt']})
        owe_them = [x for x in result if x['net'] > 0]
        owe_me = [x for x in result if x['net'] < 0]
        even = [x for x in result if x['net'] == 0]
        return jsonify({'owe_them': owe_them, 'owe_me': owe_me, 'even': even, 'total': len(result)})
    finally:
        c.close()


# ==================== 导入导出 ====================

@bp.route('/export')
def export_csv():
    eid = request.args.get('event_id')
    c = _get_db()
    try:
        if eid:
            rows = c.execute('SELECT r.*,e.name as event_name FROM records r JOIN events e ON r.event_id=e.id WHERE r.event_id=? ORDER BY r.date DESC', (eid,)).fetchall()
            en = c.execute('SELECT name FROM events WHERE id=?', (eid,)).fetchone()
            fn = ''.join(c if c.isalnum() or c in '._- ' else '_' for c in en['name']).strip() + '.csv' if en else 'records.csv'
        else:
            rows = c.execute('SELECT r.*,e.name as event_name FROM records r JOIN events e ON r.event_id=e.id ORDER BY e.id,r.date DESC').fetchall()
            fn = 'all.csv'
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(['event', 'date', 'name', 'amount', 'direction', 'note', 'batch'])
        for r in rows:
            w.writerow([r['event_name'], r['date'] or '', r['name'], r['amount'],
                         r['direction'] or '收', r['note'] or '', r['batch']])
        return Response('\ufeff' + buf.getvalue(), mimetype='text/csv',
                        headers={'Content-Disposition': f'attachment; filename={fn}'})
    finally:
        c.close()


@bp.route('/import', methods=['POST'])
def import_upload():
    if 'file' not in request.files: return jsonify({'success': False, 'error': '请选择文件'}), 400
    f = request.files['file']
    if not f.filename or not f.filename.endswith(('.xlsx', '.xls')): return jsonify({'success': False, 'error': '请上传Excel文件'}), 400
    c = _get_db()
    try:
        wb = _get_openpyxl().load_workbook(f); ws = wb.active
        eid = request.form.get('event_id')
        if not eid:
            ev = c.execute("SELECT id FROM events WHERE name=?", (ws.title,)).fetchone()
            if not ev:
                c.execute("INSERT INTO events (name,event_type) VALUES (?,?)", (ws.title, '导入')); c.commit()
                eid = c.execute("SELECT id FROM events WHERE name=?", (ws.title,)).fetchone()['id']
            else: eid = ev['id']
        else: eid = int(eid)
        lock_err = _lock_check(c, eid)
        if lock_err: return lock_err
        imp, upd, skip = 0, 0, 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) >= 4:
                dv, n, a, note = row[:4]; b = row[4] if len(row) > 4 and row[4] else 1
            elif len(row) >= 3:
                n, a, b = row[:3]; note = ''; dv = None
            else: skip += 1; continue
            if not n or not a: skip += 1; continue
            try: a = float(a)
            except (ValueError, TypeError): skip += 1; continue
            ds = dv.strftime('%Y-%m-%d') if hasattr(dv, 'strftime') else str(dv) if dv else ''
            name = str(n).strip()
            batch = int(b) if b else 1
            note_val = str(note).strip() if note else ''
            existing = c.execute('SELECT id FROM records WHERE event_id=? AND name=? AND batch=?',
                                 (eid, name, batch)).fetchone()
            direction = str(row[5]).strip() if len(row) >= 6 and row[5] else '收'
            if direction not in ('收','送'): direction = '收'
            if existing:
                c.execute('UPDATE records SET date=?,amount=?,note=?,direction=? WHERE id=?',
                          (ds, a, note_val, direction, existing['id']))
                upd += 1
            else:
                c.execute('INSERT INTO records (event_id,date,name,amount,note,batch,direction) VALUES (?,?,?,?,?,?,?)',
                          (eid, ds, name, a, note_val, batch, direction))
                imp += 1
        c.commit()
        _log(f'导入数据到事件{eid}: {imp}条新增 {upd}条更新 {skip}条跳过')
        return jsonify({'success': True, 'imported': imp, 'updated': upd, 'skipped': skip, 'event_id': eid})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        c.close()


@bp.route('/db-check')
def db_check():
    """诊断接口：测试数据库写入能力，使用 deploy token 鉴权"""
    import traceback
    result = {'steps': [], 'success': False}
    c = None
    try:
        result['steps'].append({'step': 'connect', 'status': 'ok', 'db_path': RENQING_DB_PATH})
        c = _get_db()

        # 1. 检查表结构
        cols_raw = c.execute('PRAGMA table_info(events)').fetchall()
        cols = [dict(r) for r in cols_raw]
        result['steps'].append({'step': 'pragma_table_info', 'status': 'ok', 'columns': cols})

        # 2. 检查 schema_version
        try:
            sv = c.execute('SELECT * FROM schema_version WHERE key=?', ('version',)).fetchone()
            if sv:
                # 尝试多种列名
                v = None
                for k in ('value', 'val', 'version'):
                    try: v = sv[k]; break
                    except: pass
                result['steps'].append({'step': 'schema_version', 'status': 'ok', 'version': v})
            else:
                result['steps'].append({'step': 'schema_version', 'status': 'ok', 'version': 'N/A'})
        except Exception as se:
            result['steps'].append({'step': 'schema_version', 'status': 'warn', 'msg': str(se)[:100]})

        # 3. 尝试插入测试事件
        test_name = '__DIAGNOSTIC_TEST_EVENT__'
        cols_names = [col['name'] for col in cols]
        has_ps = 'page_size' in cols_names
        if has_ps:
            c.execute("INSERT INTO events (name,event_type,event_date,page_size) VALUES (?,?,?,?)",
                      (test_name, '诊断测试', '', 12))
        else:
            c.execute("INSERT INTO events (name,event_type,event_date) VALUES (?,?,?)",
                      (test_name, '诊断测试', ''))
        result['steps'].append({'step': 'insert_test', 'status': 'ok', 'has_page_size': has_ps})

        # 4. 验证插入成功
        test_evt = c.execute('SELECT * FROM events WHERE name=?', (test_name,)).fetchone()
        if test_evt:
            result['steps'].append({'step': 'verify_insert', 'status': 'ok', 'event': dict(test_evt)})
            # 5. 删除测试事件
            c.execute('DELETE FROM events WHERE name=?', (test_name,))
            result['steps'].append({'step': 'delete_test', 'status': 'ok'})
        else:
            result['steps'].append({'step': 'verify_insert', 'status': 'fail', 'msg': '插入后未找到测试事件'})

        c.commit()
        result['success'] = True
    except Exception as e:
        result['steps'].append({'step': 'error', 'type': type(e).__name__, 'msg': str(e), 'trace': traceback.format_exc()[:500]})
        result['success'] = False
    finally:
        if c:
            try: c.rollback()
            except: pass
            try: c.close()
            except: pass
    return jsonify(result)
