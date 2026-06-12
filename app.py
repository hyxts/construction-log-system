# -*- coding: utf-8 -*-
"""
国标装修工装施工日志系统 - Flask 后端
适用于 PythonAnywhere 部署
"""
import os
import io
import json
import sqlite3
import subprocess
import traceback
import uuid
from datetime import datetime
from flask import Flask, request, jsonify, send_from_directory, send_file
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

app = Flask(__name__, static_folder='public', static_url_path='')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB 限制
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data.db')

# ==================== 全局错误处理 ====================

@app.errorhandler(Exception)
def handle_exception(e):
    """全局异常处理，返回 JSON 格式错误"""
    app.logger.error(f"Unhandled exception: {str(e)}\n{traceback.format_exc()}")
    return jsonify({'success': False, 'error': '服务器内部错误，请稍后重试'}), 500

@app.errorhandler(400)
def handle_400(e):
    return jsonify({'success': False, 'error': '请求参数错误'}), 400

@app.errorhandler(404)
def handle_404(e):
    return jsonify({'success': False, 'error': '资源不存在'}), 404

@app.errorhandler(405)
def handle_405(e):
    return jsonify({'success': False, 'error': '请求方法不允许'}), 405

# ==================== 数据库工具 ====================

def get_db():
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA busy_timeout=5000")
    return conn

def init_db():
    conn = get_db()
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS projects (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            type TEXT DEFAULT 'custom',
            company TEXT DEFAULT '',
            client TEXT DEFAULT '',
            address TEXT DEFAULT '',
            manager TEXT DEFAULT '',
            recorder TEXT DEFAULT '',
            duration INTEGER DEFAULT 0,
            start_date TEXT DEFAULT '',
            end_date TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime')),
            updated_at TEXT DEFAULT (datetime('now','localtime'))
        );

        CREATE TABLE IF NOT EXISTS tasks (
            id TEXT PRIMARY KEY,
            project_id TEXT NOT NULL,
            category TEXT DEFAULT '',
            name TEXT NOT NULL DEFAULT '',
            location TEXT DEFAULT '',
            start TEXT DEFAULT '',
            end TEXT DEFAULT '',
            team TEXT DEFAULT '',
            workers INTEGER DEFAULT 0,
            status TEXT DEFAULT 'pending',
            description TEXT DEFAULT '',
            remark TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime')),
            updated_at TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS logs (
            id TEXT PRIMARY KEY,
            project_id TEXT NOT NULL,
            project_name TEXT DEFAULT '',
            unit TEXT DEFAULT '',
            date TEXT DEFAULT '',
            weather TEXT DEFAULT '',
            temp_high TEXT DEFAULT '',
            temp_low TEXT DEFAULT '',
            wind TEXT DEFAULT '',
            location TEXT DEFAULT '',
            incident TEXT DEFAULT '',
            production_record TEXT DEFAULT '',
            tech_quality_safety TEXT DEFAULT '',
            manager TEXT DEFAULT '',
            recorder TEXT DEFAULT '',
            materials TEXT DEFAULT '',
            equipments TEXT DEFAULT '',
            daily_task_log_ids TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime')),
            updated_at TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS daily_task_logs (
            id TEXT PRIMARY KEY,
            project_id TEXT NOT NULL,
            task_id TEXT NOT NULL,
            log_date TEXT DEFAULT '',
            content TEXT DEFAULT '',
            weather TEXT DEFAULT '',
            team TEXT DEFAULT '',
            worker_count INTEGER DEFAULT 0,
            materials TEXT DEFAULT '',
            equipments TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE CASCADE,
            FOREIGN KEY (task_id) REFERENCES tasks(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS teams (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL DEFAULT '',
            leader TEXT DEFAULT '',
            phone TEXT DEFAULT '',
            specialty TEXT DEFAULT '',
            worker_count INTEGER DEFAULT 0,
            remark TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime')),
            updated_at TEXT DEFAULT (datetime('now','localtime'))
        );

        CREATE TABLE IF NOT EXISTS materials (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL DEFAULT '',
            spec TEXT DEFAULT '',
            unit TEXT DEFAULT '',
            quantity REAL DEFAULT 0,
            min_quantity REAL DEFAULT 0,
            supplier TEXT DEFAULT '',
            status TEXT DEFAULT 'in_stock',
            remark TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime')),
            updated_at TEXT DEFAULT (datetime('now','localtime'))
        );

        CREATE TABLE IF NOT EXISTS equipments (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL DEFAULT '',
            model TEXT DEFAULT '',
            count INTEGER DEFAULT 1,
            status TEXT DEFAULT 'normal',
            remark TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime')),
            updated_at TEXT DEFAULT (datetime('now','localtime'))
        );

        CREATE TABLE IF NOT EXISTS acceptances (
            id TEXT PRIMARY KEY,
            project_id TEXT NOT NULL,
            task_id TEXT DEFAULT '',
            rework_task_id TEXT DEFAULT '',
            acceptance_type TEXT DEFAULT 'normal',
            name TEXT NOT NULL DEFAULT '',
            location TEXT DEFAULT '',
            unit TEXT DEFAULT '',
            basis TEXT DEFAULT '',
            design_qty REAL DEFAULT 0,
            actual_qty REAL DEFAULT 0,
            unit_price REAL DEFAULT 0,
            total_price REAL DEFAULT 0,
            calc_formula TEXT DEFAULT '',
            quantity_type TEXT DEFAULT '',
            status TEXT DEFAULT 'pending',
            date TEXT DEFAULT '',
            remark TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime')),
            updated_at TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS project_photos (
            id TEXT PRIMARY KEY,
            project_id TEXT NOT NULL,
            task_id TEXT DEFAULT '',
            category TEXT DEFAULT '',
            filename TEXT DEFAULT '',
            data TEXT DEFAULT '',
            thumbnail TEXT DEFAULT '',
            file_size INTEGER DEFAULT 0,
            mime_type TEXT DEFAULT 'image/jpeg',
            description TEXT DEFAULT '',
            taken_at TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS misc_finance (
            id TEXT PRIMARY KEY,
            project_ref TEXT DEFAULT '',
            flow_type TEXT DEFAULT 'expense',
            source TEXT DEFAULT 'personal',
            category TEXT DEFAULT '零星材料',
            amount REAL DEFAULT 0,
            description TEXT DEFAULT '',
            date TEXT DEFAULT '',
            carryover_mode TEXT DEFAULT '',
            carryover_month TEXT DEFAULT '',
            carryover_ref_id TEXT DEFAULT '',
            status TEXT DEFAULT 'active',
            created_at TEXT DEFAULT (datetime('now','localtime')),
            updated_at TEXT DEFAULT (datetime('now','localtime'))
        );

        CREATE TABLE IF NOT EXISTS reworks (
            id TEXT PRIMARY KEY,
            project_id TEXT NOT NULL,
            task_id TEXT DEFAULT '',
            name TEXT NOT NULL DEFAULT '',
            rework_qty REAL DEFAULT 0,
            unit TEXT DEFAULT 'm²',
            reason TEXT DEFAULT '',
            status TEXT DEFAULT 'pending',
            start_date TEXT DEFAULT '',
            end_date TEXT DEFAULT '',
            remark TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime')),
            updated_at TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE CASCADE
        );

        CREATE INDEX IF NOT EXISTS idx_tasks_project ON tasks(project_id);
        CREATE INDEX IF NOT EXISTS idx_logs_project ON logs(project_id);
        CREATE INDEX IF NOT EXISTS idx_logs_date ON logs(date);
        CREATE INDEX IF NOT EXISTS idx_daily_task_logs_project ON daily_task_logs(project_id);
        CREATE INDEX IF NOT EXISTS idx_daily_task_logs_task ON daily_task_logs(task_id);
        CREATE INDEX IF NOT EXISTS idx_daily_task_logs_date ON daily_task_logs(log_date);
        CREATE INDEX IF NOT EXISTS idx_acceptances_project ON acceptances(project_id);
        CREATE INDEX IF NOT EXISTS idx_photos_project ON project_photos(project_id);
        CREATE INDEX IF NOT EXISTS idx_finance_date ON misc_finance(date);
        CREATE INDEX IF NOT EXISTS idx_finance_source ON misc_finance(source);
        CREATE INDEX IF NOT EXISTS idx_reworks_project ON reworks(project_id);
        CREATE INDEX IF NOT EXISTS idx_reworks_task ON reworks(task_id);
    ''')
    # Migration: add missing columns (silently skip if exist)
    for table, col, col_def in [
        ('projects', 'duration', 'INTEGER DEFAULT 0'),
        ('projects', 'start_date', "TEXT DEFAULT ''"),
        ('projects', 'end_date', "TEXT DEFAULT ''"),
        ('tasks', 'workers', 'INTEGER DEFAULT 0'),
        ('logs', 'materials', "TEXT DEFAULT ''"),
        ('logs', 'equipments', "TEXT DEFAULT ''"),
        ('logs', 'daily_task_log_ids', "TEXT DEFAULT ''"),
        ('daily_task_logs', 'weather', "TEXT DEFAULT ''"),
        ('daily_task_logs', 'team', "TEXT DEFAULT ''"),
        ('daily_task_logs', 'worker_count', 'INTEGER DEFAULT 0'),
        ('daily_task_logs', 'materials', "TEXT DEFAULT ''"),
        ('daily_task_logs', 'equipments', "TEXT DEFAULT ''"),
        ('acceptances', 'basis', "TEXT DEFAULT ''"),
        ('acceptances', 'calc_formula', "TEXT DEFAULT ''"),
        ('acceptances', 'task_id', "TEXT DEFAULT ''"),
        ('acceptances', 'rework_task_id', "TEXT DEFAULT ''"),
        ('acceptances', 'acceptance_type', "TEXT DEFAULT 'normal'"),
        ('acceptances', 'quantity_type', "TEXT DEFAULT ''"),
        ('acceptances', 'unit_price', 'REAL DEFAULT 0'),
        ('acceptances', 'total_price', 'REAL DEFAULT 0'),
        ('materials', 'min_quantity', 'REAL DEFAULT 0'),
    ]:
        try:
            conn.execute(f"ALTER TABLE {table} ADD COLUMN {col} {col_def}")
        except Exception as mig_err:
            app.logger.warning(f"Migration add column skipped: {table}.{col} - {mig_err}")

    # ---- Migration: 移除 teams/materials/equipments 的 project_id，实现多项目通用 ----
    for tbl in ['teams', 'materials', 'equipments']:
        try:
            cols = [row[1] for row in conn.execute(f"PRAGMA table_info({tbl})").fetchall()]
            if 'project_id' in cols:
                # 创建新表（无 project_id）
                col_defs = []
                for row in conn.execute(f"PRAGMA table_info({tbl})").fetchall():
                    if row[1] == 'project_id':
                        continue
                    col_defs.append(f"{row[1]} {row[2]}" + (f" DEFAULT {row[4]}" if row[4] is not None else ""))
                conn.execute(f"CREATE TABLE {tbl}_new ({','.join(col_defs)})")
                # 复制数据
                non_pid_cols = [c for c in cols if c != 'project_id']
                conn.execute(f"INSERT INTO {tbl}_new ({','.join(non_pid_cols)}) SELECT {','.join(non_pid_cols)} FROM {tbl}")
                # 替换
                conn.execute(f"DROP TABLE {tbl}")
                conn.execute(f"ALTER TABLE {tbl}_new RENAME TO {tbl}")
                app.logger.info(f"Migration: removed project_id from {tbl}")
        except Exception as mig_err:
            app.logger.warning(f"Migration {tbl} project_id removal skipped: {mig_err}")

    conn.commit()
    conn.close()

def dict_from_row(row):
    if row is None:
        return None
    return dict(row)

def dicts_from_rows(rows):
    return [dict(row) for row in rows]

# ==================== 工程 CRUD ====================

@app.route('/api/projects', methods=['GET'])
def get_projects():
    conn = get_db()
    projects = dicts_from_rows(conn.execute('SELECT * FROM projects ORDER BY created_at DESC').fetchall())
    # 批量查询所有任务，避免 N+1 问题
    all_tasks = dicts_from_rows(conn.execute('SELECT * FROM tasks ORDER BY start, id').fetchall())
    task_map = {}
    for t in all_tasks:
        task_map.setdefault(t['project_id'], []).append(t)
    for p in projects:
        p['tasks'] = task_map.get(p['id'], [])
    conn.close()
    return jsonify({'success': True, 'data': projects})

@app.route('/api/projects/<pid>', methods=['GET'])
def get_project(pid):
    conn = get_db()
    project = dict_from_row(conn.execute('SELECT * FROM projects WHERE id = ?', (pid,)).fetchone())
    if not project:
        conn.close()
        return jsonify({'success': False, 'error': '工程不存在'}), 404
    tasks = dicts_from_rows(
        conn.execute('SELECT * FROM tasks WHERE project_id = ? ORDER BY start, id', (pid,)).fetchall()
    )
    project['tasks'] = tasks
    conn.close()
    return jsonify({'success': True, 'data': project})

@app.route('/api/projects', methods=['POST'])
def create_project():
    data = (request.get_json(silent=True) or {})
    name = data.get('name', '')
    if not name:
        return jsonify({'success': False, 'error': '工程名称不能为空'}), 400
    pid = data.get('id') or str(int(datetime.now().timestamp() * 1000))
    conn = get_db()
    conn.execute(
        'INSERT INTO projects (id, name, type, company, client, address, manager, recorder, duration, start_date, end_date) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
        (pid, name, data.get('type', 'custom'), data.get('company', ''),
         data.get('client', ''), data.get('address', ''), data.get('manager', ''), data.get('recorder', ''),
         data.get('duration', 0), data.get('start_date', ''), data.get('end_date', ''))
    )
    conn.commit()
    project = dict_from_row(conn.execute('SELECT * FROM projects WHERE id = ?', (pid,)).fetchone())
    project['tasks'] = []
    conn.close()
    return jsonify({'success': True, 'data': project})

@app.route('/api/projects/<pid>', methods=['PUT'])
def update_project(pid):
    data = (request.get_json(silent=True) or {})
    conn = get_db()
    conn.execute(
        "UPDATE projects SET name=?, type=?, company=?, client=?, address=?, manager=?, recorder=?, duration=?, start_date=?, end_date=?, updated_at=datetime('now','localtime') WHERE id=?",
        (data.get('name', ''), data.get('type', 'custom'), data.get('company', ''),
         data.get('client', ''), data.get('address', ''), data.get('manager', ''), data.get('recorder', ''),
         data.get('duration', 0), data.get('start_date', ''), data.get('end_date', ''), pid)
    )
    conn.commit()
    project = dict_from_row(conn.execute('SELECT * FROM projects WHERE id = ?', (pid,)).fetchone())
    if not project:
        conn.close()
        return jsonify({'success': False, 'error': '工程不存在'}), 404
    tasks = dicts_from_rows(
        conn.execute('SELECT * FROM tasks WHERE project_id = ? ORDER BY start, id', (pid,)).fetchall()
    )
    project['tasks'] = tasks
    conn.close()
    return jsonify({'success': True, 'data': project})

@app.route('/api/projects/<pid>', methods=['DELETE'])
def delete_project(pid):
    conn = get_db()
    conn.execute('DELETE FROM daily_task_logs WHERE project_id = ?', (pid,))
    conn.execute('DELETE FROM logs WHERE project_id = ?', (pid,))
    conn.execute('DELETE FROM acceptances WHERE project_id = ?', (pid,))
    conn.execute('DELETE FROM project_photos WHERE project_id = ?', (pid,))
    conn.execute('DELETE FROM reworks WHERE project_id = ?', (pid,))
    conn.execute('DELETE FROM tasks WHERE project_id = ?', (pid,))
    conn.execute('DELETE FROM projects WHERE id = ?', (pid,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

# ==================== 任务 CRUD ====================

@app.route('/api/tasks/project/<project_id>', methods=['GET'])
def get_tasks(project_id):
    conn = get_db()
    tasks = dicts_from_rows(
        conn.execute('SELECT * FROM tasks WHERE project_id = ? ORDER BY start, id', (project_id,)).fetchall()
    )
    conn.close()
    return jsonify({'success': True, 'data': tasks})

@app.route('/api/tasks', methods=['POST'])
def create_task():
    data = (request.get_json(silent=True) or {})
    project_id = data.get('project_id', '')
    if not project_id:
        return jsonify({'success': False, 'error': 'project_id 不能为空'}), 400
    tid = data.get('id') or str(int(datetime.now().timestamp() * 1000))
    conn = get_db()
    conn.execute(
        '''INSERT INTO tasks (id, project_id, category, name, location, start, end, team, workers, status, description, remark)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
        (tid, project_id, data.get('category', ''), data.get('name', ''),
         data.get('location', ''), data.get('start', ''), data.get('end', ''),
         data.get('team', ''), data.get('workers', 0), data.get('status', 'pending'),
         data.get('description', ''), data.get('remark', ''))
    )
    conn.commit()
    task = dict_from_row(conn.execute('SELECT * FROM tasks WHERE id = ?', (tid,)).fetchone())
    conn.close()
    return jsonify({'success': True, 'data': task})

@app.route('/api/tasks/<tid>', methods=['PUT'])
def update_task(tid):
    data = (request.get_json(silent=True) or {})
    conn = get_db()
    conn.execute(
        "UPDATE tasks SET category=?, name=?, location=?, start=?, end=?, team=?, workers=?, status=?, description=?, remark=?, updated_at=datetime('now','localtime') WHERE id=?",
        (data.get('category', ''), data.get('name', ''), data.get('location', ''),
         data.get('start', ''), data.get('end', ''), data.get('team', ''),
         data.get('workers', 0), data.get('status', 'pending'),
         data.get('description', ''), data.get('remark', ''), tid)
    )
    conn.commit()
    task = dict_from_row(conn.execute('SELECT * FROM tasks WHERE id = ?', (tid,)).fetchone())
    conn.close()
    if not task:
        return jsonify({'success': False, 'error': '任务不存在'}), 404
    return jsonify({'success': True, 'data': task})

@app.route('/api/tasks/<tid>', methods=['DELETE'])
def delete_task(tid):
    conn = get_db()
    conn.execute('DELETE FROM tasks WHERE id = ?', (tid,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/tasks/<tid>/status', methods=['PATCH'])
def update_task_status(tid):
    data = (request.get_json(silent=True) or {})
    status = data.get('status', 'pending')
    conn = get_db()
    conn.execute("UPDATE tasks SET status=?, updated_at=datetime('now','localtime') WHERE id=?", (status, tid))
    conn.commit()
    task = dict_from_row(conn.execute('SELECT * FROM tasks WHERE id = ?', (tid,)).fetchone())
    conn.close()
    return jsonify({'success': True, 'data': task})

@app.route('/api/tasks/batch', methods=['POST'])
def batch_create_tasks():
    data = (request.get_json(silent=True) or {})
    tasks = data.get('tasks', [])
    if not tasks:
        return jsonify({'success': False, 'error': 'tasks 不能为空'}), 400
    conn = get_db()
    count = 0
    for t in tasks:
        tid = t.get('id') or str(int(datetime.now().timestamp() * 1000)) + str(count)
        conn.execute(
            '''INSERT INTO tasks (id, project_id, category, name, location, start, end, team, workers, status, description, remark)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
            (tid, t.get('project_id', ''), t.get('category', ''), t.get('name', ''),
             t.get('location', ''), t.get('start', ''), t.get('end', ''),
             t.get('team', ''), t.get('workers', 0), t.get('status', 'pending'),
             t.get('description', ''), t.get('remark', ''))
        )
        count += 1
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'data': {'count': count}})

# ==================== 日志 CRUD ====================

@app.route('/api/logs/project/<project_id>', methods=['GET'])
def get_logs(project_id):
    conn = get_db()
    logs = dicts_from_rows(
        conn.execute('SELECT * FROM logs WHERE project_id = ? ORDER BY date DESC, created_at DESC', (project_id,)).fetchall()
    )
    conn.close()
    return jsonify({'success': True, 'data': logs})

@app.route('/api/logs/<lid>', methods=['GET'])
def get_log(lid):
    conn = get_db()
    log = dict_from_row(conn.execute('SELECT * FROM logs WHERE id = ?', (lid,)).fetchone())
    conn.close()
    if not log:
        return jsonify({'success': False, 'error': '日志不存在'}), 404
    return jsonify({'success': True, 'data': log})

@app.route('/api/logs', methods=['POST'])
def create_log():
    data = (request.get_json(silent=True) or {})
    project_id = data.get('project_id', '')
    if not project_id:
        return jsonify({'success': False, 'error': 'project_id 不能为空'}), 400
    lid = data.get('id') or str(int(datetime.now().timestamp() * 1000))
    conn = get_db()
    conn.execute(
        '''INSERT INTO logs (id, project_id, project_name, unit, date, weather,
            temp_high, temp_low, wind, location, incident,
            production_record, tech_quality_safety, manager, recorder, materials, equipments, daily_task_log_ids)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
        (lid, project_id, data.get('project_name', ''), data.get('unit', ''),
         data.get('date', ''), data.get('weather', ''),
         data.get('temp_high', ''), data.get('temp_low', ''), data.get('wind', ''),
         data.get('location', ''), data.get('incident', ''),
         data.get('production_record', ''), data.get('tech_quality_safety', ''),
         data.get('manager', ''), data.get('recorder', ''),
         json.dumps(data.get('materials', []), ensure_ascii=False) if data.get('materials') else '',
         json.dumps(data.get('equipments', []), ensure_ascii=False) if data.get('equipments') else '',
         data.get('daily_task_log_ids', ''))
    )
    conn.commit()
    log = dict_from_row(conn.execute('SELECT * FROM logs WHERE id = ?', (lid,)).fetchone())
    conn.close()
    return jsonify({'success': True, 'data': log})

@app.route('/api/logs/<lid>', methods=['PUT'])
def update_log(lid):
    data = (request.get_json(silent=True) or {})
    conn = get_db()
    conn.execute(
        "UPDATE logs SET project_name=?, unit=?, date=?, weather=?, temp_high=?, temp_low=?, wind=?, location=?, incident=?, production_record=?, tech_quality_safety=?, manager=?, recorder=?, materials=?, equipments=?, daily_task_log_ids=?, updated_at=datetime('now','localtime') WHERE id=?",
        (data.get('project_name', ''), data.get('unit', ''), data.get('date', ''),
         data.get('weather', ''), data.get('temp_high', ''), data.get('temp_low', ''),
         data.get('wind', ''), data.get('location', ''), data.get('incident', ''),
         data.get('production_record', ''), data.get('tech_quality_safety', ''),
         data.get('manager', ''), data.get('recorder', ''),
         json.dumps(data.get('materials', []), ensure_ascii=False) if data.get('materials') else '',
         json.dumps(data.get('equipments', []), ensure_ascii=False) if data.get('equipments') else '',
         data.get('daily_task_log_ids', ''), lid)
    )
    conn.commit()
    log = dict_from_row(conn.execute('SELECT * FROM logs WHERE id = ?', (lid,)).fetchone())
    conn.close()
    if not log:
        return jsonify({'success': False, 'error': '日志不存在'}), 404
    return jsonify({'success': True, 'data': log})

@app.route('/api/logs/<lid>', methods=['DELETE'])
def delete_log(lid):
    conn = get_db()
    conn.execute('DELETE FROM logs WHERE id = ?', (lid,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

# ==================== 每日任务施工日志 CRUD ====================

@app.route('/api/daily-task-logs/project/<project_id>', methods=['GET'])
def get_daily_task_logs(project_id):
    """获取某工程所有每日任务施工日志"""
    conn = get_db()
    logs = dicts_from_rows(
        conn.execute('SELECT * FROM daily_task_logs WHERE project_id = ? ORDER BY log_date ASC', (project_id,)).fetchall()
    )
    conn.close()
    return jsonify({'success': True, 'data': logs})

@app.route('/api/daily-task-logs/task/<task_id>', methods=['GET'])
def get_task_daily_logs(task_id):
    """获取某任务的所有每日施工日志"""
    conn = get_db()
    logs = dicts_from_rows(
        conn.execute('SELECT * FROM daily_task_logs WHERE task_id = ? ORDER BY log_date ASC', (task_id,)).fetchall()
    )
    conn.close()
    return jsonify({'success': True, 'data': logs})

@app.route('/api/daily-task-logs', methods=['POST'])
def create_daily_task_log():
    """创建或更新每日任务施工日志（同一任务+日期唯一）"""
    data = (request.get_json(silent=True) or {})
    task_id = data.get('task_id', '')
    log_date = data.get('log_date', '')
    project_id = data.get('project_id', '')
    if not task_id or not log_date:
        return jsonify({'success': False, 'error': 'task_id 和 log_date 不能为空'}), 400

    conn = get_db()
    # 查找是否已存在
    existing = conn.execute(
        'SELECT id FROM daily_task_logs WHERE task_id = ? AND log_date = ?', (task_id, log_date)
    ).fetchone()

    if existing:
        # 更新
        conn.execute(
            "UPDATE daily_task_logs SET content=?, weather=?, team=?, worker_count=?, materials=?, equipments=?, created_at=datetime('now','localtime') WHERE id=?",
            (data.get('content', ''), data.get('weather', ''), data.get('team', ''),
             data.get('worker_count', 0),
             json.dumps(data.get('materials', []), ensure_ascii=False) if data.get('materials') else '',
             json.dumps(data.get('equipments', []), ensure_ascii=False) if data.get('equipments') else '',
             existing['id'])
        )
        lid = existing['id']
    else:
        # 创建
        lid = str(uuid.uuid4())
        conn.execute(
            '''INSERT INTO daily_task_logs (id, project_id, task_id, log_date, content, weather, team, worker_count, materials, equipments)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
            (lid, project_id, task_id, log_date, data.get('content', ''),
             data.get('weather', ''), data.get('team', ''), data.get('worker_count', 0),
             json.dumps(data.get('materials', []), ensure_ascii=False) if data.get('materials') else '',
             json.dumps(data.get('equipments', []), ensure_ascii=False) if data.get('equipments') else '')
        )
    conn.commit()
    log = dict_from_row(conn.execute('SELECT * FROM daily_task_logs WHERE id = ?', (lid,)).fetchone())
    conn.close()
    return jsonify({'success': True, 'data': log})

@app.route('/api/daily-task-logs/<lid>', methods=['DELETE'])
def delete_daily_task_log(lid):
    """删除每日任务施工日志"""
    conn = get_db()
    conn.execute('DELETE FROM daily_task_logs WHERE id = ?', (lid,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

# ==================== 人员/班组/材料/设备 管理 (personnel) ====================

@app.route('/api/personnel/<resource_type>/project/<project_id>', methods=['GET'])
@app.route('/api/personnel/<resource_type>/all', methods=['GET'])
def get_personnel(resource_type, project_id=None):
    """获取班组/材料/设备列表（全项目通用）"""
    table_map = {'teams': 'teams', 'materials': 'materials', 'equipments': 'equipments'}
    table = table_map.get(resource_type)
    if not table:
        return jsonify({'success': False, 'error': '无效资源类型'}), 400
    conn = get_db()
    items = dicts_from_rows(
        conn.execute(f'SELECT * FROM {table} ORDER BY created_at DESC').fetchall()
    )
    conn.close()
    return jsonify({'success': True, 'data': items})

@app.route('/api/personnel/<resource_type>', methods=['POST'])
def create_personnel(resource_type):
    """创建班组/材料/设备"""
    table_map = {'teams': 'teams', 'materials': 'materials', 'equipments': 'equipments'}
    table = table_map.get(resource_type)
    if not table:
        return jsonify({'success': False, 'error': '无效资源类型'}), 400
    data = (request.get_json(silent=True) or {})
    rid = data.get('id') or str(int(datetime.now().timestamp() * 1000))
    conn = get_db()
    if table == 'teams':
        conn.execute(
            f'INSERT INTO {table} (id, name, leader, phone, specialty, worker_count, remark) VALUES (?, ?, ?, ?, ?, ?, ?)',
            (rid, data.get('name', ''), data.get('leader', ''), data.get('phone', ''),
             data.get('specialty', ''), data.get('worker_count', 0), data.get('remark', ''))
        )
    elif table == 'materials':
        conn.execute(
            f'INSERT INTO {table} (id, name, spec, unit, quantity, min_quantity, supplier, status, remark) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)',
            (rid, data.get('name', ''), data.get('spec', ''), data.get('unit', ''),
             data.get('quantity', 0), data.get('min_quantity', 0), data.get('supplier', ''),
             data.get('status', 'in_stock'), data.get('remark', ''))
        )
    elif table == 'equipments':
        conn.execute(
            f'INSERT INTO {table} (id, name, model, count, status, remark) VALUES (?, ?, ?, ?, ?, ?)',
            (rid, data.get('name', ''), data.get('model', ''),
             data.get('count', 1), data.get('status', 'normal'), data.get('remark', ''))
        )
    conn.commit()
    item = dict_from_row(conn.execute(f'SELECT * FROM {table} WHERE id = ?', (rid,)).fetchone())
    conn.close()
    return jsonify({'success': True, 'data': item})

@app.route('/api/personnel/<resource_type>/<rid>', methods=['PUT'])
def update_personnel(resource_type, rid):
    """更新班组/材料/设备"""
    table_map = {'teams': 'teams', 'materials': 'materials', 'equipments': 'equipments'}
    table = table_map.get(resource_type)
    if not table:
        return jsonify({'success': False, 'error': '无效资源类型'}), 400
    data = (request.get_json(silent=True) or {})
    conn = get_db()
    if table == 'teams':
        conn.execute(
            f"UPDATE {table} SET name=?, leader=?, phone=?, specialty=?, worker_count=?, remark=?, updated_at=datetime('now','localtime') WHERE id=?",
            (data.get('name', ''), data.get('leader', ''), data.get('phone', ''),
             data.get('specialty', ''), data.get('worker_count', 0), data.get('remark', ''), rid)
        )
    elif table == 'materials':
        conn.execute(
            f"UPDATE {table} SET name=?, spec=?, unit=?, quantity=?, min_quantity=?, supplier=?, status=?, remark=?, updated_at=datetime('now','localtime') WHERE id=?",
            (data.get('name', ''), data.get('spec', ''), data.get('unit', ''),
             data.get('quantity', 0), data.get('min_quantity', 0), data.get('supplier', ''),
             data.get('status', 'in_stock'), data.get('remark', ''), rid)
        )
    elif table == 'equipments':
        conn.execute(
            f"UPDATE {table} SET name=?, model=?, count=?, status=?, remark=?, updated_at=datetime('now','localtime') WHERE id=?",
            (data.get('name', ''), data.get('model', ''), data.get('count', 1),
             data.get('status', 'normal'), data.get('remark', ''), rid)
        )
    conn.commit()
    item = dict_from_row(conn.execute(f'SELECT * FROM {table} WHERE id = ?', (rid,)).fetchone())
    conn.close()
    return jsonify({'success': True, 'data': item})

@app.route('/api/personnel/<resource_type>/<rid>', methods=['DELETE'])
def delete_personnel(resource_type, rid):
    """删除班组/材料/设备"""
    table_map = {'teams': 'teams', 'materials': 'materials', 'equipments': 'equipments'}
    table = table_map.get(resource_type)
    if not table:
        return jsonify({'success': False, 'error': '无效资源类型'}), 400
    conn = get_db()
    conn.execute(f'DELETE FROM {table} WHERE id = ?', (rid,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

# ==================== 收方管理 (acceptances) ====================

@app.route('/api/acceptances/project/<project_id>', methods=['GET'])
def get_acceptances(project_id):
    """获取某工程所有收方记录"""
    conn = get_db()
    items = dicts_from_rows(
        conn.execute('SELECT * FROM acceptances WHERE project_id = ? ORDER BY date DESC, created_at DESC', (project_id,)).fetchall()
    )
    conn.close()
    return jsonify({'success': True, 'data': items})

@app.route('/api/acceptances', methods=['POST'])
def create_acceptance():
    """创建收方记录"""
    data = (request.get_json(silent=True) or {})
    project_id = data.get('project_id', '')
    if not project_id:
        return jsonify({'success': False, 'error': 'project_id 不能为空'}), 400
    aid = data.get('id') or str(int(datetime.now().timestamp() * 1000))
    conn = get_db()
    conn.execute(
        '''INSERT INTO acceptances (id, project_id, task_id, rework_task_id, acceptance_type,
            name, location, unit, basis, design_qty, actual_qty, unit_price, total_price,
            calc_formula, quantity_type, status, date, remark)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
        (aid, project_id, data.get('task_id', ''), data.get('rework_task_id', ''),
         data.get('acceptance_type', 'normal'), data.get('name', ''),
         data.get('location', ''), data.get('unit', ''), data.get('basis', ''),
         data.get('design_qty', 0), data.get('actual_qty', 0),
         data.get('unit_price', 0), data.get('total_price', 0),
         data.get('calc_formula', ''), data.get('quantity_type', ''),
         data.get('status', 'pending'), data.get('date', ''), data.get('remark', ''))
    )
    conn.commit()
    item = dict_from_row(conn.execute('SELECT * FROM acceptances WHERE id = ?', (aid,)).fetchone())
    conn.close()
    return jsonify({'success': True, 'data': item})

@app.route('/api/acceptances/<aid>', methods=['PUT'])
def update_acceptance(aid):
    """更新收方记录"""
    data = (request.get_json(silent=True) or {})
    conn = get_db()
    conn.execute(
        "UPDATE acceptances SET task_id=?, rework_task_id=?, acceptance_type=?, name=?, location=?, unit=?, basis=?, design_qty=?, actual_qty=?, unit_price=?, total_price=?, calc_formula=?, quantity_type=?, status=?, date=?, remark=?, updated_at=datetime('now','localtime') WHERE id=?",
        (data.get('task_id', ''), data.get('rework_task_id', ''), data.get('acceptance_type', 'normal'),
         data.get('name', ''), data.get('location', ''), data.get('unit', ''),
         data.get('basis', ''), data.get('design_qty', 0), data.get('actual_qty', 0),
         data.get('unit_price', 0), data.get('total_price', 0),
         data.get('calc_formula', ''), data.get('quantity_type', ''),
         data.get('status', 'pending'), data.get('date', ''), data.get('remark', ''), aid)
    )
    conn.commit()
    item = dict_from_row(conn.execute('SELECT * FROM acceptances WHERE id = ?', (aid,)).fetchone())
    conn.close()
    if not item:
        return jsonify({'success': False, 'error': '收方记录不存在'}), 404
    return jsonify({'success': True, 'data': item})

@app.route('/api/acceptances/<aid>/status', methods=['PATCH'])
def update_acceptance_status(aid):
    """仅更新收方状态（确认/反确认），不覆盖其他字段"""
    data = (request.get_json(silent=True) or {})
    status = data.get('status', 'confirmed')
    conn = get_db()
    conn.execute("UPDATE acceptances SET status=?, updated_at=datetime('now','localtime') WHERE id=?", (status, aid))
    conn.commit()
    item = dict_from_row(conn.execute('SELECT * FROM acceptances WHERE id = ?', (aid,)).fetchone())
    conn.close()
    if not item:
        return jsonify({'success': False, 'error': '收方记录不存在'}), 404
    return jsonify({'success': True, 'data': item})

@app.route('/api/acceptances/<aid>', methods=['DELETE'])
def delete_acceptance(aid):
    """删除收方记录"""
    conn = get_db()
    conn.execute('DELETE FROM acceptances WHERE id = ?', (aid,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

# ==================== 工程影像管理 (project_photos) ====================

@app.route('/api/photos/project/<project_id>', methods=['GET'])
def get_photos(project_id):
    """获取某工程所有照片（不含完整图片数据，仅缩略图+元信息）"""
    conn = get_db()
    photos = dicts_from_rows(
        conn.execute('SELECT id, project_id, task_id, category, filename, thumbnail, file_size, mime_type, description, taken_at, created_at FROM project_photos WHERE project_id = ? ORDER BY created_at DESC', (project_id,)).fetchall()
    )
    conn.close()
    return jsonify({'success': True, 'data': photos})

@app.route('/api/photos/<photo_id>', methods=['GET'])
def get_photo_full(photo_id):
    """获取单张照片完整数据（含原图）"""
    conn = get_db()
    photo = dict_from_row(conn.execute('SELECT * FROM project_photos WHERE id = ?', (photo_id,)).fetchone())
    conn.close()
    if not photo:
        return jsonify({'success': False, 'error': '照片不存在'}), 404
    return jsonify({'success': True, 'data': photo})

@app.route('/api/photos', methods=['POST'])
def create_photo():
    """上传照片（base64格式）"""
    data = (request.get_json(silent=True) or {})
    project_id = data.get('project_id', '')
    if not project_id:
        return jsonify({'success': False, 'error': 'project_id 不能为空'}), 400
    img_data = data.get('data', '')
    if not img_data:
        return jsonify({'success': False, 'error': '图片数据不能为空'}), 400

    pid = data.get('id') or str(int(datetime.now().timestamp() * 1000))
    # 缩略图：前端预览时通过 CSS 缩放，这里不截断 base64 数据以保证图片完整
    thumbnail = data.get('thumbnail', '') or img_data

    conn = get_db()
    conn.execute(
        '''INSERT INTO project_photos (id, project_id, task_id, category, filename, data, thumbnail, file_size, mime_type, description, taken_at)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
        (pid, project_id, data.get('task_id', ''), data.get('category', ''),
         data.get('filename', 'photo.jpg'), img_data, thumbnail,
         data.get('file_size', 0), data.get('mime_type', 'image/jpeg'),
         data.get('description', ''), data.get('taken_at', ''))
    )
    conn.commit()
    photo = dict_from_row(
        conn.execute('SELECT id, project_id, task_id, category, filename, thumbnail, file_size, mime_type, description, taken_at, created_at FROM project_photos WHERE id = ?', (pid,)).fetchone()
    )
    conn.close()
    return jsonify({'success': True, 'data': photo})

@app.route('/api/photos/<photo_id>', methods=['PUT'])
def update_photo(photo_id):
    """更新照片信息（分类、描述等）"""
    data = (request.get_json(silent=True) or {})
    conn = get_db()
    conn.execute(
        "UPDATE project_photos SET task_id=?, category=?, description=?, taken_at=? WHERE id=?",
        (data.get('task_id', ''), data.get('category', ''), data.get('description', ''), data.get('taken_at', ''), photo_id)
    )
    conn.commit()
    photo = dict_from_row(
        conn.execute('SELECT id, project_id, task_id, category, filename, thumbnail, file_size, mime_type, description, taken_at, created_at FROM project_photos WHERE id = ?', (photo_id,)).fetchone()
    )
    conn.close()
    return jsonify({'success': True, 'data': photo})

@app.route('/api/photos/<photo_id>', methods=['DELETE'])
def delete_photo(photo_id):
    """删除照片"""
    conn = get_db()
    conn.execute('DELETE FROM project_photos WHERE id = ?', (photo_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

# ==================== 零星财务记账 (misc_finance) ====================

@app.route('/api/finance', methods=['GET'])
def get_finance_list():
    """获取零星财务流水（支持筛选）"""
    conn = get_db()
    source = request.args.get('source', '')
    category = request.args.get('category', '')
    status = request.args.get('status', '')
    month = request.args.get('month', '')  # YYYY-MM
    search = request.args.get('search', '')

    sql = 'SELECT * FROM misc_finance WHERE 1=1'
    params = []
    if source:
        sql += ' AND source = ?'
        params.append(source)
    if category:
        sql += ' AND category = ?'
        params.append(category)
    if status:
        sql += ' AND status = ?'
        params.append(status)
    if month:
        sql += " AND substr(date,1,7) = ?"
        params.append(month)
    if search:
        sql += ' AND (description LIKE ? OR project_ref LIKE ? OR category LIKE ?)'
        like = '%' + search + '%'
        params.extend([like, like, like])
    sql += ' ORDER BY date DESC, created_at DESC'

    items = dicts_from_rows(conn.execute(sql, params).fetchall())
    conn.close()
    return jsonify({'success': True, 'data': items})

@app.route('/api/finance/<fid>', methods=['GET'])
def get_finance(fid):
    """获取单条财务记录"""
    conn = get_db()
    item = dict_from_row(conn.execute('SELECT * FROM misc_finance WHERE id = ?', (fid,)).fetchone())
    conn.close()
    if not item:
        return jsonify({'success': False, 'error': '记录不存在'}), 404
    return jsonify({'success': True, 'data': item})

@app.route('/api/finance', methods=['POST'])
def create_finance():
    """创建零星财务流水"""
    data = (request.get_json(silent=True) or {})
    fid = data.get('id') or str(int(datetime.now().timestamp() * 1000))
    conn = get_db()
    conn.execute(
        '''INSERT INTO misc_finance (id, project_ref, flow_type, source, category, amount, description, date, carryover_mode, carryover_month, carryover_ref_id, status)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
        (fid, data.get('project_ref', ''), data.get('flow_type', 'expense'),
         data.get('source', 'personal'), data.get('category', '零星材料'),
         data.get('amount', 0), data.get('description', ''), data.get('date', ''),
         data.get('carryover_mode', ''), data.get('carryover_month', ''),
         data.get('carryover_ref_id', ''), data.get('status', 'active'))
    )
    conn.commit()
    item = dict_from_row(conn.execute('SELECT * FROM misc_finance WHERE id = ?', (fid,)).fetchone())
    conn.close()
    return jsonify({'success': True, 'data': item})

@app.route('/api/finance/<fid>', methods=['PUT'])
def update_finance(fid):
    """更新零星财务流水"""
    data = (request.get_json(silent=True) or {})
    conn = get_db()
    conn.execute(
        "UPDATE misc_finance SET project_ref=?, flow_type=?, source=?, category=?, amount=?, description=?, date=?, carryover_mode=?, carryover_month=?, carryover_ref_id=?, status=?, updated_at=datetime('now','localtime') WHERE id=?",
        (data.get('project_ref', ''), data.get('flow_type', 'expense'),
         data.get('source', 'personal'), data.get('category', '零星材料'),
         data.get('amount', 0), data.get('description', ''), data.get('date', ''),
         data.get('carryover_mode', ''), data.get('carryover_month', ''),
         data.get('carryover_ref_id', ''), data.get('status', 'active'), fid)
    )
    conn.commit()
    item = dict_from_row(conn.execute('SELECT * FROM misc_finance WHERE id = ?', (fid,)).fetchone())
    conn.close()
    return jsonify({'success': True, 'data': item})

@app.route('/api/finance/<fid>', methods=['DELETE'])
def delete_finance(fid):
    """删除零星财务记录"""
    conn = get_db()
    conn.execute('DELETE FROM misc_finance WHERE id = ?', (fid,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/finance/carryover', methods=['POST'])
def carryover_finance():
    """批量结转财务记录"""
    data = (request.get_json(silent=True) or {})
    ids = data.get('ids', [])
    mode = data.get('mode', 'monthly')  # monthly | project
    month = data.get('month', '')  # for monthly: YYYY-MM
    ref_id = data.get('ref_id', '')  # for project: project_id
    if not ids:
        return jsonify({'success': False, 'error': '请选择要结转的记录'}), 400

    conn = get_db()
    if mode == 'monthly':
        conn.execute(
            "UPDATE misc_finance SET status='carried_over', carryover_mode='monthly', carryover_month=?, updated_at=datetime('now','localtime') WHERE id IN ({})".format(','.join('?'*len(ids))),
            [month] + ids
        )
    else:
        conn.execute(
            "UPDATE misc_finance SET status='carried_over', carryover_mode='project', carryover_ref_id=?, updated_at=datetime('now','localtime') WHERE id IN ({})".format(','.join('?'*len(ids))),
            [ref_id] + ids
        )
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'data': {'count': len(ids)}})

@app.route('/api/finance/stats', methods=['GET'])
def get_finance_stats():
    """财务统计摘要"""
    conn = get_db()
    # 按来源统计
    source_stats = dicts_from_rows(conn.execute(
        "SELECT source, SUM(CASE WHEN flow_type='expense' THEN amount ELSE 0 END) as total_expense, SUM(CASE WHEN flow_type='income' THEN amount ELSE 0 END) as total_income, COUNT(*) as cnt FROM misc_finance WHERE status='active' GROUP BY source"
    ).fetchall())
    # 按类别统计
    cat_stats = dicts_from_rows(conn.execute(
        "SELECT category, SUM(amount) as total, COUNT(*) as cnt FROM misc_finance WHERE status='active' AND flow_type='expense' GROUP BY category ORDER BY total DESC"
    ).fetchall())
    # 按月统计
    month_stats = dicts_from_rows(conn.execute(
        "SELECT substr(date,1,7) as month, SUM(amount) as total, COUNT(*) as cnt FROM misc_finance WHERE status='active' GROUP BY month ORDER BY month DESC LIMIT 12"
    ).fetchall())
    conn.close()
    return jsonify({'success': True, 'data': {'source_stats': source_stats, 'category_stats': cat_stats, 'month_stats': month_stats}})

# ==================== 数据报表 ====================

@app.route('/api/reports/finance', methods=['GET'])
def report_finance():
    """财务报表 - 按来源/类别/月份/项目多维度统计"""
    conn = get_db()
    project_id = request.args.get('project_id', '')
    # 总支出/总收入
    totals = dict_from_row(conn.execute(
        "SELECT COALESCE(SUM(CASE WHEN flow_type='expense' THEN amount ELSE 0 END),0) as total_expense, COALESCE(SUM(CASE WHEN flow_type='income' THEN amount ELSE 0 END),0) as total_income, COUNT(*) as total_count FROM misc_finance WHERE status='active'"
        + (" AND project_ref=?" if project_id else ""),
        (project_id,) if project_id else ()
    ).fetchone())
    # 按来源分组
    by_source = dicts_from_rows(conn.execute(
        "SELECT source, SUM(amount) as total, COUNT(*) as cnt, SUM(CASE WHEN flow_type='expense' THEN amount ELSE 0 END) as expense, SUM(CASE WHEN flow_type='income' THEN amount ELSE 0 END) as income FROM misc_finance WHERE status='active'"
        + (" AND project_ref=?" if project_id else "") +
        " GROUP BY source",
        (project_id,) if project_id else ()
    ).fetchall())
    # 按类别（支出）分组
    by_category = dicts_from_rows(conn.execute(
        "SELECT category, SUM(amount) as total, COUNT(*) as cnt FROM misc_finance WHERE status='active' AND flow_type='expense'"
        + (" AND project_ref=?" if project_id else "") +
        " GROUP BY category ORDER BY total DESC",
        (project_id,) if project_id else ()
    ).fetchall())
    # 按月份分组
    by_month = dicts_from_rows(conn.execute(
        "SELECT substr(date,1,7) as month, SUM(CASE WHEN flow_type='expense' THEN amount ELSE 0 END) as expense, SUM(CASE WHEN flow_type='income' THEN amount ELSE 0 END) as income, COUNT(*) as cnt FROM misc_finance WHERE status='active'"
        + (" AND project_ref=?" if project_id else "") +
        " GROUP BY month ORDER BY month DESC LIMIT 24",
        (project_id,) if project_id else ()
    ).fetchall())
    # 按项目分组
    by_project = dicts_from_rows(conn.execute(
        "SELECT mf.project_ref, COALESCE(p.name, mf.project_ref) as project_name, SUM(CASE WHEN mf.flow_type='expense' THEN mf.amount ELSE 0 END) as expense, SUM(CASE WHEN mf.flow_type='income' THEN mf.amount ELSE 0 END) as income, COUNT(*) as cnt FROM misc_finance mf LEFT JOIN projects p ON mf.project_ref=p.id WHERE mf.status='active' GROUP BY mf.project_ref ORDER BY expense DESC"
    ).fetchall())
    conn.close()
    return jsonify({'success': True, 'data': {
        'totals': totals, 'by_source': by_source, 'by_category': by_category,
        'by_month': by_month, 'by_project': by_project
    }})

@app.route('/api/reports/acceptance', methods=['GET'])
def report_acceptance():
    """收方报表 - 按项目/任务/类别多维度统计"""
    conn = get_db()
    project_id = request.args.get('project_id', '')
    # 总览
    totals = dict_from_row(conn.execute(
        "SELECT COUNT(*) as total_count, COALESCE(SUM(design_qty),0) as total_design, COALESCE(SUM(actual_qty),0) as total_actual, COALESCE(SUM(total_price),0) as total_price, COUNT(CASE WHEN status='confirmed' THEN 1 END) as confirmed_count, COUNT(CASE WHEN status='pending' THEN 1 END) as pending_count FROM acceptances"
        + (" WHERE project_id=?" if project_id else ""),
        (project_id,) if project_id else ()
    ).fetchone())
    # 按项目分组
    by_project = dicts_from_rows(conn.execute(
        "SELECT a.project_id, COALESCE(p.name, a.project_id) as project_name, COUNT(*) as cnt, COALESCE(SUM(a.design_qty),0) as total_design, COALESCE(SUM(a.actual_qty),0) as total_actual, COALESCE(SUM(a.total_price),0) as total_price FROM acceptances a LEFT JOIN projects p ON a.project_id=p.id"
        + (" WHERE a.project_id=?" if project_id else "") +
        " GROUP BY a.project_id ORDER BY cnt DESC",
        (project_id,) if project_id else ()
    ).fetchall())
    # 按计量单位统计
    by_unit = dicts_from_rows(conn.execute(
        "SELECT unit, COUNT(*) as cnt, COALESCE(SUM(design_qty),0) as total_design, COALESCE(SUM(actual_qty),0) as total_actual FROM acceptances"
        + (" WHERE project_id=?" if project_id else "") +
        " GROUP BY unit ORDER BY cnt DESC",
        (project_id,) if project_id else ()
    ).fetchall())
    # 按日期分组
    by_date = dicts_from_rows(conn.execute(
        "SELECT substr(date,1,7) as month, COUNT(*) as cnt, COALESCE(SUM(design_qty),0) as total_design, COALESCE(SUM(actual_qty),0) as total_actual FROM acceptances"
        + (" WHERE project_id=?" if project_id else "") +
        " GROUP BY month ORDER BY month DESC LIMIT 24",
        (project_id,) if project_id else ()
    ).fetchall())
    # 明细列表
    items = dicts_from_rows(conn.execute(
        "SELECT a.*, p.name as project_name FROM acceptances a LEFT JOIN projects p ON a.project_id=p.id"
        + (" WHERE a.project_id=?" if project_id else "") +
        " ORDER BY a.date DESC, a.created_at DESC LIMIT 500",
        (project_id,) if project_id else ()
    ).fetchall())
    conn.close()
    return jsonify({'success': True, 'data': {
        'totals': totals, 'by_project': by_project, 'by_unit': by_unit,
        'by_date': by_date, 'items': items
    }})

@app.route('/api/reports/personnel', methods=['GET'])
def report_personnel():
    """人员报表 - 班组/任务/工时多维度统计"""
    conn = get_db()
    project_id = request.args.get('project_id', '')
    # 班组统计
    team_stats = dicts_from_rows(conn.execute(
        "SELECT t.id, t.name, t.worker_count as workers, t.specialty, COUNT(tk.id) as task_count, SUM(CASE WHEN tk.status='completed' THEN 1 ELSE 0 END) as completed_tasks FROM teams t LEFT JOIN tasks tk ON tk.team=t.name" +
        (" AND tk.project_id=?" if project_id else "") +
        " GROUP BY t.id ORDER BY task_count DESC",
        (project_id,) if project_id else ()
    ).fetchall())
    # 任务状态统计
    task_status = dict_from_row(conn.execute(
        "SELECT COUNT(*) as total, COALESCE(SUM(CASE WHEN status='pending' THEN 1 ELSE 0 END),0) as pending, COALESCE(SUM(CASE WHEN status='in_progress' THEN 1 ELSE 0 END),0) as in_progress, COALESCE(SUM(CASE WHEN status='completed' THEN 1 ELSE 0 END),0) as completed FROM tasks"
        + (" WHERE project_id=?" if project_id else ""),
        (project_id,) if project_id else ()
    ).fetchone())
    # 按分类统计任务
    task_by_category = dicts_from_rows(conn.execute(
        "SELECT category, COUNT(*) as cnt, COALESCE(SUM(CASE WHEN status='completed' THEN 1 ELSE 0 END),0) as completed FROM tasks"
        + (" WHERE project_id=?" if project_id else "") +
        " GROUP BY category ORDER BY cnt DESC",
        (project_id,) if project_id else ()
    ).fetchall())
    # 按班组名称统计任务
    task_by_team = dicts_from_rows(conn.execute(
        "SELECT team, COUNT(*) as cnt, COALESCE(SUM(CASE WHEN status='completed' THEN 1 ELSE 0 END),0) as completed, COALESCE(SUM(CASE WHEN status='in_progress' THEN 1 ELSE 0 END),0) as in_progress FROM tasks WHERE team!=''"
        + (" AND project_id=?" if project_id else "") +
        " GROUP BY team ORDER BY cnt DESC",
        (project_id,) if project_id else ()
    ).fetchall())
    # 返工统计
    rework_stats = dict_from_row(conn.execute(
        "SELECT COUNT(*) as total, COALESCE(SUM(CASE WHEN status='pending' THEN 1 ELSE 0 END),0) as pending, COALESCE(SUM(CASE WHEN status='in_progress' THEN 1 ELSE 0 END),0) as in_progress, COALESCE(SUM(CASE WHEN status='completed' THEN 1 ELSE 0 END),0) as completed, COALESCE(SUM(rework_qty),0) as total_qty FROM reworks"
        + (" WHERE project_id=?" if project_id else ""),
        (project_id,) if project_id else ()
    ).fetchone())
    # 总工人数
    total_workers = dict_from_row(conn.execute(
        "SELECT COALESCE(SUM(worker_count),0) as total FROM teams"
    ).fetchone())
    conn.close()
    return jsonify({'success': True, 'data': {
        'team_stats': team_stats, 'task_status': task_status,
        'task_by_category': task_by_category, 'task_by_team': task_by_team,
        'rework_stats': rework_stats, 'total_workers': total_workers
    }})

@app.route('/api/reports/combined', methods=['GET'])
def report_combined():
    """综合报表 - 汇总财务+收方+人员三维度"""
    conn = get_db()
    # 财务汇总
    fin = dict_from_row(conn.execute(
        "SELECT COALESCE(SUM(CASE WHEN flow_type='expense' THEN amount ELSE 0 END),0) as total_expense, COALESCE(SUM(CASE WHEN flow_type='income' THEN amount ELSE 0 END),0) as total_income, COUNT(*) as count FROM misc_finance WHERE status='active'"
    ).fetchone())
    # 收方汇总
    acc = dict_from_row(conn.execute(
        "SELECT COUNT(*) as count, COALESCE(SUM(design_qty),0) as total_design, COALESCE(SUM(actual_qty),0) as total_actual, COALESCE(SUM(total_price),0) as total_price, COUNT(CASE WHEN status='confirmed' THEN 1 END) as confirmed FROM acceptances"
    ).fetchone())
    # 任务汇总
    task = dict_from_row(conn.execute(
        "SELECT COUNT(*) as total, COALESCE(SUM(CASE WHEN status='completed' THEN 1 ELSE 0 END),0) as completed, COALESCE(SUM(CASE WHEN status='in_progress' THEN 1 ELSE 0 END),0) as in_progress FROM tasks"
    ).fetchone())
    # 班组汇总
    team_sum = dict_from_row(conn.execute(
        "SELECT COUNT(*) as count, COALESCE(SUM(worker_count),0) as total_workers FROM teams"
    ).fetchone())
    # 返工汇总
    rw = dict_from_row(conn.execute(
        "SELECT COUNT(*) as total, COALESCE(SUM(CASE WHEN status='completed' THEN 1 ELSE 0 END),0) as completed, COALESCE(SUM(CASE WHEN status!='completed' THEN 1 ELSE 0 END),0) as pending FROM reworks"
    ).fetchone())
    # 按项目汇总 - 使用批量查询替代相关子查询，大幅提升性能
    by_project = dicts_from_rows(conn.execute(
        "SELECT id, name, '' as status FROM projects ORDER BY created_at DESC"
    ).fetchall())
    # 任务统计
    task_rows = conn.execute(
        "SELECT project_id, COUNT(*) as cnt, SUM(CASE WHEN status='completed' THEN 1 ELSE 0 END) as done FROM tasks GROUP BY project_id"
    ).fetchall()
    task_stats = {r['project_id']: (r['cnt'], r['done'] or 0) for r in task_rows}
    # 财务统计
    fin_rows = conn.execute(
        "SELECT project_ref, SUM(CASE WHEN flow_type='expense' AND status='active' THEN amount ELSE 0 END) as expense, SUM(CASE WHEN flow_type='income' AND status='active' THEN amount ELSE 0 END) as income FROM misc_finance GROUP BY project_ref"
    ).fetchall()
    fin_stats = {r['project_ref']: (r['expense'] or 0, r['income'] or 0) for r in fin_rows}
    # 验收统计
    acc_rows = conn.execute(
        "SELECT project_id, COUNT(*) as cnt, COALESCE(SUM(total_price),0) as total FROM acceptances GROUP BY project_id"
    ).fetchall()
    acc_stats = {r['project_id']: (r['cnt'], r['total'] or 0) for r in acc_rows}
    for p in by_project:
        pid = p['id']
        tc, td = task_stats.get(pid, (0, 0))
        ex, inc = fin_stats.get(pid, (0, 0))
        ac, at = acc_stats.get(pid, (0, 0))
        p['task_count'] = tc
        p['task_done'] = td
        p['expense'] = ex
        p['income'] = inc
        p['acceptance_count'] = ac
        p['acceptance_total'] = at
    # 工程总数
    proj_count = conn.execute("SELECT COUNT(*) as cnt FROM projects").fetchone()['cnt']
    conn.close()
    return jsonify({'success': True, 'data': {
        'finance': fin, 'acceptance': acc, 'tasks': task,
        'teams': team_sum, 'reworks': rw, 'by_project': by_project,
        'project_count': proj_count
    }})

# ==================== Excel 导出 ====================

# 通用样式
HEADER_FONT = Font(name='Microsoft YaHei', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
TITLE_FONT = Font(name='Microsoft YaHei', bold=True, size=14)
SUBTITLE_FONT = Font(name='Microsoft YaHei', size=10, color='666666')
CELL_FONT = Font(name='Microsoft YaHei', size=10)
CELL_ALIGN = Alignment(vertical='center', wrap_text=True)
NUM_ALIGN = Alignment(horizontal='right', vertical='center')
THIN_BORDER = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB')
)

def _make_excel_response(wb, filename):
    """将 Workbook 转为 Flask 下载响应"""
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

def _style_header_row(ws, row, cols):
    """设置表头行样式"""
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

def _style_data_cell(ws, row, col, is_number=False):
    """设置数据单元格样式"""
    cell = ws.cell(row=row, column=col)
    cell.font = CELL_FONT
    cell.alignment = NUM_ALIGN if is_number else CELL_ALIGN
    cell.border = THIN_BORDER

def _auto_width(ws, min_width=8, max_width=40):
    """自动调整列宽"""
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                # 中文字符约占2个宽度单位
                line_len = 0
                for ch in str(val):
                    line_len += 2 if ord(ch) > 127 else 1
                max_len = max(max_len, line_len)
        width = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

def _add_title_sheet(wb, ws, title, subtitle=''):
    """写入标题和副标题"""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column or 6)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ws.max_column or 6)
        sub_cell = ws.cell(row=2, column=1, value=subtitle)
        sub_cell.font = SUBTITLE_FONT
        sub_cell.alignment = Alignment(horizontal='center', vertical='center')


# ---------- 财务报表 Excel ----------

@app.route('/api/reports/finance/excel', methods=['GET'])
def report_finance_excel():
    conn = get_db()
    project_id = request.args.get('project_id', '')
    totals = dict_from_row(conn.execute(
        "SELECT COALESCE(SUM(CASE WHEN flow_type='expense' THEN amount ELSE 0 END),0) as total_expense, COALESCE(SUM(CASE WHEN flow_type='income' THEN amount ELSE 0 END),0) as total_income, COUNT(*) as total_count FROM misc_finance WHERE status='active'"
        + (" AND project_ref=?" if project_id else ""),
        (project_id,) if project_id else ()
    ).fetchone())
    by_source = dicts_from_rows(conn.execute(
        "SELECT source, SUM(amount) as total, COUNT(*) as cnt, SUM(CASE WHEN flow_type='expense' THEN amount ELSE 0 END) as expense, SUM(CASE WHEN flow_type='income' THEN amount ELSE 0 END) as income FROM misc_finance WHERE status='active'"
        + (" AND project_ref=?" if project_id else "") +
        " GROUP BY source",
        (project_id,) if project_id else ()
    ).fetchall())
    by_category = dicts_from_rows(conn.execute(
        "SELECT category, SUM(amount) as total, COUNT(*) as cnt FROM misc_finance WHERE status='active' AND flow_type='expense'"
        + (" AND project_ref=?" if project_id else "") +
        " GROUP BY category ORDER BY total DESC",
        (project_id,) if project_id else ()
    ).fetchall())
    by_month = dicts_from_rows(conn.execute(
        "SELECT substr(date,1,7) as month, SUM(CASE WHEN flow_type='expense' THEN amount ELSE 0 END) as expense, SUM(CASE WHEN flow_type='income' THEN amount ELSE 0 END) as income, COUNT(*) as cnt FROM misc_finance WHERE status='active'"
        + (" AND project_ref=?" if project_id else "") +
        " GROUP BY month ORDER BY month DESC LIMIT 24",
        (project_id,) if project_id else ()
    ).fetchall())
    conn.close()

    wb = Workbook()
    # Sheet 1: 总览
    ws = wb.active
    ws.title = '总览与来源'
    _add_title_sheet(wb, ws, '财务报表', f'导出时间：{datetime.now().strftime("%Y-%m-%d %H:%M")}')
    curr_row = 4
    headers = ['指标', '金额/数量']
    for ci, h in enumerate(headers, 1):
        ws.cell(row=curr_row, column=ci, value=h)
    _style_header_row(ws, curr_row, len(headers))
    balance = (totals.get('total_income', 0) or 0) - (totals.get('total_expense', 0) or 0)
    overview_data = [
        ('总支出', f"¥{(totals.get('total_expense', 0) or 0):.2f}"),
        ('总收入', f"¥{(totals.get('total_income', 0) or 0):.2f}"),
        ('盈余/赤字', f"¥{balance:.2f} ({'盈余' if balance >= 0 else '赤字'})"),
        ('总记录数', str(totals.get('total_count', 0) or 0)),
    ]
    for i, (label, val) in enumerate(overview_data):
        r = curr_row + 1 + i
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=val)
        _style_data_cell(ws, r, 1)
        _style_data_cell(ws, r, 2, True)

    # 按资金来源
    if by_source:
        curr_row = curr_row + len(overview_data) + 3
        src_headers = ['来源', '支出', '收入', '合计', '笔数']
        for ci, h in enumerate(src_headers, 1):
            ws.cell(row=curr_row, column=ci, value=h)
        _style_header_row(ws, curr_row, len(src_headers))
        for i, s in enumerate(by_source):
            r = curr_row + 1 + i
            ws.cell(row=r, column=1, value='公司财务' if s.get('source') == 'company' else '个人')
            ws.cell(row=r, column=2, value=float(s.get('expense', 0) or 0))
            ws.cell(row=r, column=3, value=float(s.get('income', 0) or 0))
            ws.cell(row=r, column=4, value=float(s.get('total', 0) or 0))
            ws.cell(row=r, column=5, value=int(s.get('cnt', 0) or 0))
            for c in range(1, 6):
                _style_data_cell(ws, r, c, c >= 2)
    _auto_width(ws)

    # Sheet 2: 支出类别
    if by_category:
        ws2 = wb.create_sheet('支出类别')
        _add_title_sheet(wb, ws2, '支出类别排行')
        curr_row = 4
        cat_headers = ['类别', '金额', '笔数']
        for ci, h in enumerate(cat_headers, 1):
            ws2.cell(row=curr_row, column=ci, value=h)
        _style_header_row(ws2, curr_row, len(cat_headers))
        for i, c in enumerate(by_category):
            r = curr_row + 1 + i
            ws2.cell(row=r, column=1, value=c.get('category', ''))
            ws2.cell(row=r, column=2, value=float(c.get('total', 0) or 0))
            ws2.cell(row=r, column=3, value=int(c.get('cnt', 0) or 0))
            _style_data_cell(ws2, r, 1)
            _style_data_cell(ws2, r, 2, True)
            _style_data_cell(ws2, r, 3, True)
        _auto_width(ws2)

    # Sheet 3: 月度趋势
    if by_month:
        ws3 = wb.create_sheet('月度趋势')
        _add_title_sheet(wb, ws3, '月度收支趋势')
        curr_row = 4
        mon_headers = ['月份', '支出', '收入', '笔数']
        for ci, h in enumerate(mon_headers, 1):
            ws3.cell(row=curr_row, column=ci, value=h)
        _style_header_row(ws3, curr_row, len(mon_headers))
        for i, m in enumerate(by_month):
            r = curr_row + 1 + i
            ws3.cell(row=r, column=1, value=m.get('month', ''))
            ws3.cell(row=r, column=2, value=float(m.get('expense', 0) or 0))
            ws3.cell(row=r, column=3, value=float(m.get('income', 0) or 0))
            ws3.cell(row=r, column=4, value=int(m.get('cnt', 0) or 0))
            _style_data_cell(ws3, r, 1)
            _style_data_cell(ws3, r, 2, True)
            _style_data_cell(ws3, r, 3, True)
            _style_data_cell(ws3, r, 4, True)
        _auto_width(ws3)

    return _make_excel_response(wb, f'财务报表_{datetime.now().strftime("%Y%m%d")}.xlsx')


# ---------- 收方报表 Excel ----------

@app.route('/api/reports/acceptance/excel', methods=['GET'])
def report_acceptance_excel():
    conn = get_db()
    project_id = request.args.get('project_id', '')
    totals = dict_from_row(conn.execute(
        "SELECT COUNT(*) as total_count, COALESCE(SUM(design_qty),0) as total_design, COALESCE(SUM(actual_qty),0) as total_actual, COALESCE(SUM(total_price),0) as total_price, COUNT(CASE WHEN status='confirmed' THEN 1 END) as confirmed_count, COUNT(CASE WHEN status='pending' THEN 1 END) as pending_count FROM acceptances"
        + (" WHERE project_id=?" if project_id else ""),
        (project_id,) if project_id else ()
    ).fetchone())
    items = dicts_from_rows(conn.execute(
        "SELECT a.*, p.name as project_name FROM acceptances a LEFT JOIN projects p ON a.project_id=p.id"
        + (" WHERE a.project_id=?" if project_id else "") +
        " ORDER BY a.date DESC, a.created_at DESC LIMIT 500",
        (project_id,) if project_id else ()
    ).fetchall())
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = '收方报表'
    _add_title_sheet(wb, ws, '收方报表', f'导出时间：{datetime.now().strftime("%Y-%m-%d %H:%M")}')

    # 总览
    curr_row = 4
    ov_headers = ['指标', '数值']
    for ci, h in enumerate(ov_headers, 1):
        ws.cell(row=curr_row, column=ci, value=h)
    _style_header_row(ws, curr_row, len(ov_headers))
    overview = [
        ('收方记录数', str(totals.get('total_count', 0) or 0)),
        ('累计总价', f"¥{(totals.get('total_price', 0) or 0):.2f}"),
        ('设计总量', str(totals.get('total_design', 0) or 0)),
        ('实际总量', str(totals.get('total_actual', 0) or 0)),
        ('已确认', str(totals.get('confirmed_count', 0) or 0)),
        ('待确认', str(totals.get('pending_count', 0) or 0)),
    ]
    for i, (label, val) in enumerate(overview):
        r = curr_row + 1 + i
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=val)
        _style_data_cell(ws, r, 1)
        _style_data_cell(ws, r, 2, True)

    # 明细
    if items:
        curr_row = curr_row + len(overview) + 3
        det_headers = ['名称', '项目', '日期', '设计量', '实际量', '单位', '单价', '总价', '状态']
        for ci, h in enumerate(det_headers, 1):
            ws.cell(row=curr_row, column=ci, value=h)
        _style_header_row(ws, curr_row, len(det_headers))
        for i, itm in enumerate(items):
            r = curr_row + 1 + i
            ws.cell(row=r, column=1, value=itm.get('name', ''))
            ws.cell(row=r, column=2, value=itm.get('project_name', ''))
            ws.cell(row=r, column=3, value=(itm.get('date', '') or '')[:10])
            ws.cell(row=r, column=4, value=float(itm.get('design_qty', 0) or 0))
            ws.cell(row=r, column=5, value=float(itm.get('actual_qty', 0) or 0))
            ws.cell(row=r, column=6, value=itm.get('unit', ''))
            ws.cell(row=r, column=7, value=float(itm.get('unit_price', 0) or 0))
            ws.cell(row=r, column=8, value=float(itm.get('total_price', 0) or 0))
            ws.cell(row=r, column=9, value='已确认' if itm.get('status') == 'confirmed' else '待确认')
            for c in range(1, 10):
                _style_data_cell(ws, r, c, c in (4, 5, 7, 8))
    _auto_width(ws)

    return _make_excel_response(wb, f'收方报表_{datetime.now().strftime("%Y%m%d")}.xlsx')


# ---------- 人员报表 Excel ----------

@app.route('/api/reports/personnel/excel', methods=['GET'])
def report_personnel_excel():
    conn = get_db()
    project_id = request.args.get('project_id', '')
    team_stats = dicts_from_rows(conn.execute(
        "SELECT t.id, t.name, t.worker_count as workers, t.specialty, COUNT(tk.id) as task_count, SUM(CASE WHEN tk.status='completed' THEN 1 ELSE 0 END) as completed_tasks FROM teams t LEFT JOIN tasks tk ON tk.team=t.name" +
        (" AND tk.project_id=?" if project_id else "") +
        " GROUP BY t.id ORDER BY task_count DESC",
        (project_id,) if project_id else ()
    ).fetchall())
    task_status = dict_from_row(conn.execute(
        "SELECT COUNT(*) as total, COALESCE(SUM(CASE WHEN status='pending' THEN 1 ELSE 0 END),0) as pending, COALESCE(SUM(CASE WHEN status='in_progress' THEN 1 ELSE 0 END),0) as in_progress, COALESCE(SUM(CASE WHEN status='completed' THEN 1 ELSE 0 END),0) as completed FROM tasks"
        + (" WHERE project_id=?" if project_id else ""),
        (project_id,) if project_id else ()
    ).fetchone())
    task_by_category = dicts_from_rows(conn.execute(
        "SELECT category, COUNT(*) as cnt, COALESCE(SUM(CASE WHEN status='completed' THEN 1 ELSE 0 END),0) as completed FROM tasks"
        + (" WHERE project_id=?" if project_id else "") +
        " GROUP BY category ORDER BY cnt DESC",
        (project_id,) if project_id else ()
    ).fetchall())
    conn.close()

    wb = Workbook()
    # Sheet 1: 总览
    ws = wb.active
    ws.title = '人员总览'
    _add_title_sheet(wb, ws, '人员报表', f'导出时间：{datetime.now().strftime("%Y-%m-%d %H:%M")}')
    curr_row = 4
    ov_headers = ['指标', '数值']
    for ci, h in enumerate(ov_headers, 1):
        ws.cell(row=curr_row, column=ci, value=h)
    _style_header_row(ws, curr_row, len(ov_headers))
    ts = task_status or {}
    total = ts.get('total', 0) or 0
    pct = round((ts.get('completed', 0) or 0) / total * 100) if total else 0
    overview = [
        ('任务总数', str(total)),
        (f'已完成 ({pct}%)', str(ts.get('completed', 0) or 0)),
        ('进行中', str(ts.get('in_progress', 0) or 0)),
        ('待开始', str(ts.get('pending', 0) or 0)),
        ('班组数', str(len(team_stats))),
    ]
    for i, (label, val) in enumerate(overview):
        r = curr_row + 1 + i
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=val)
        _style_data_cell(ws, r, 1)
        _style_data_cell(ws, r, 2, True)

    # 班组统计
    if team_stats:
        curr_row = curr_row + len(overview) + 3
        tm_headers = ['班组', '人数', '任务数', '已完成', '完成率', '专业']
        for ci, h in enumerate(tm_headers, 1):
            ws.cell(row=curr_row, column=ci, value=h)
        _style_header_row(ws, curr_row, len(tm_headers))
        for i, tm in enumerate(team_stats):
            r = curr_row + 1 + i
            tc = tm.get('task_count', 0) or 0
            rate = round((tm.get('completed_tasks', 0) or 0) / tc * 100) if tc else 0
            ws.cell(row=r, column=1, value=tm.get('name', ''))
            ws.cell(row=r, column=2, value=int(tm.get('workers', 0) or 0))
            ws.cell(row=r, column=3, value=tc)
            ws.cell(row=r, column=4, value=int(tm.get('completed_tasks', 0) or 0))
            ws.cell(row=r, column=5, value=f'{rate}%')
            ws.cell(row=r, column=6, value=tm.get('specialty', ''))
            for c in range(1, 7):
                _style_data_cell(ws, r, c, c in (2, 3, 4))
        _auto_width(ws)

    # Sheet 2: 工序分类
    if task_by_category:
        ws2 = wb.create_sheet('工序分类')
        _add_title_sheet(wb, ws2, '工序分类统计')
        curr_row = 4
        cat_headers = ['分类', '任务数', '已完成', '完成率']
        for ci, h in enumerate(cat_headers, 1):
            ws2.cell(row=curr_row, column=ci, value=h)
        _style_header_row(ws2, curr_row, len(cat_headers))
        for i, c in enumerate(task_by_category):
            r = curr_row + 1 + i
            cnt = c.get('cnt', 0) or 0
            comp = c.get('completed', 0) or 0
            rate = round(comp / cnt * 100) if cnt else 0
            ws2.cell(row=r, column=1, value=c.get('category', ''))
            ws2.cell(row=r, column=2, value=cnt)
            ws2.cell(row=r, column=3, value=comp)
            ws2.cell(row=r, column=4, value=f'{rate}%')
            for ci2 in range(1, 5):
                _style_data_cell(ws2, r, ci2, ci2 >= 2)
        _auto_width(ws2)

    return _make_excel_response(wb, f'人员报表_{datetime.now().strftime("%Y%m%d")}.xlsx')


# ---------- 综合报表 Excel ----------

@app.route('/api/reports/combined/excel', methods=['GET'])
def report_combined_excel():
    conn = get_db()
    fin = dict_from_row(conn.execute(
        "SELECT COALESCE(SUM(CASE WHEN flow_type='expense' THEN amount ELSE 0 END),0) as total_expense, COALESCE(SUM(CASE WHEN flow_type='income' THEN amount ELSE 0 END),0) as total_income, COUNT(*) as count FROM misc_finance WHERE status='active'"
    ).fetchone())
    acc = dict_from_row(conn.execute(
        "SELECT COUNT(*) as count, COALESCE(SUM(design_qty),0) as total_design, COALESCE(SUM(actual_qty),0) as total_actual, COALESCE(SUM(total_price),0) as total_price, COUNT(CASE WHEN status='confirmed' THEN 1 END) as confirmed FROM acceptances"
    ).fetchone())
    task = dict_from_row(conn.execute(
        "SELECT COUNT(*) as total, COALESCE(SUM(CASE WHEN status='completed' THEN 1 ELSE 0 END),0) as completed, COALESCE(SUM(CASE WHEN status='in_progress' THEN 1 ELSE 0 END),0) as in_progress FROM tasks"
    ).fetchone())
    team_sum = dict_from_row(conn.execute(
        "SELECT COUNT(*) as count, COALESCE(SUM(worker_count),0) as total_workers FROM teams"
    ).fetchone())
    rw = dict_from_row(conn.execute(
        "SELECT COUNT(*) as total, COALESCE(SUM(CASE WHEN status='completed' THEN 1 ELSE 0 END),0) as completed, COALESCE(SUM(CASE WHEN status!='completed' THEN 1 ELSE 0 END),0) as pending FROM reworks"
    ).fetchone())
    by_project = dicts_from_rows(conn.execute(
        "SELECT p.id, p.name, (SELECT COUNT(*) FROM tasks t WHERE t.project_id=p.id) as task_count, (SELECT COUNT(*) FROM tasks t WHERE t.project_id=p.id AND t.status='completed') as task_done, (SELECT COALESCE(SUM(amount),0) FROM misc_finance mf WHERE mf.project_ref=p.id AND mf.flow_type='expense' AND mf.status='active') as expense, (SELECT COALESCE(SUM(amount),0) FROM misc_finance mf2 WHERE mf2.project_ref=p.id AND mf2.flow_type='income' AND mf2.status='active') as income, (SELECT COUNT(*) FROM acceptances a WHERE a.project_id=p.id) as acceptance_count, (SELECT COALESCE(SUM(a.total_price),0) FROM acceptances a WHERE a.project_id=p.id) as acceptance_total FROM projects p ORDER BY p.created_at DESC"
    ).fetchall())
    proj_count = conn.execute("SELECT COUNT(*) as cnt FROM projects").fetchone()['cnt']
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = '综合报表'
    _add_title_sheet(wb, ws, '综合报表', f'涵盖 {proj_count} 个工程 · 导出时间：{datetime.now().strftime("%Y-%m-%d %H:%M")}')

    # 概览卡片
    curr_row = 4
    ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=8)
    sec_cell = ws.cell(row=curr_row, column=1, value='财务与收方概览')
    sec_cell.font = Font(name='Microsoft YaHei', bold=True, size=11, color='2563EB')

    curr_row += 1
    ov_headers = ['总收入', '总支出', '收支结余', '收方总价', '任务总数', '已完成', '人员总数', '返工记录']
    for ci, h in enumerate(ov_headers, 1):
        ws.cell(row=curr_row, column=ci, value=h)
    _style_header_row(ws, curr_row, len(ov_headers))
    balance = (fin.get('total_income', 0) or 0) - (fin.get('total_expense', 0) or 0)
    task_comp = task.get('completed', 0) or 0
    ov_vals = [
        f"¥{(fin.get('total_income', 0) or 0):.2f}",
        f"¥{(fin.get('total_expense', 0) or 0):.2f}",
        f"¥{balance:.2f}",
        f"¥{(acc.get('total_price', 0) or 0):.2f}",
        str(task.get('total', 0) or 0),
        str(task_comp),
        str(team_sum.get('total_workers', 0) or 0),
        str(rw.get('total', 0) or 0),
    ]
    ov_row = curr_row + 1
    for ci, val in enumerate(ov_vals, 1):
        ws.cell(row=ov_row, column=ci, value=val)
        _style_data_cell(ws, ov_row, ci, True)

    # 各工程明细
    if by_project:
        curr_row = ov_row + 3
        ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=8)
        sec_cell2 = ws.cell(row=curr_row, column=1, value='各工程明细')
        sec_cell2.font = Font(name='Microsoft YaHei', bold=True, size=11, color='2563EB')

        curr_row += 1
        pj_headers = ['工程名称', '任务数', '已完成', '收入', '支出', '结余', '收方记录', '收方总价']
        for ci, h in enumerate(pj_headers, 1):
            ws.cell(row=curr_row, column=ci, value=h)
        _style_header_row(ws, curr_row, len(pj_headers))
        for i, p in enumerate(by_project):
            r = curr_row + 1 + i
            p_balance = (p.get('income', 0) or 0) - (p.get('expense', 0) or 0)
            ws.cell(row=r, column=1, value=p.get('name', ''))
            ws.cell(row=r, column=2, value=int(p.get('task_count', 0) or 0))
            ws.cell(row=r, column=3, value=int(p.get('task_done', 0) or 0))
            ws.cell(row=r, column=4, value=float(p.get('income', 0) or 0))
            ws.cell(row=r, column=5, value=float(p.get('expense', 0) or 0))
            ws.cell(row=r, column=6, value=p_balance)
            ws.cell(row=r, column=7, value=int(p.get('acceptance_count', 0) or 0))
            ws.cell(row=r, column=8, value=float(p.get('acceptance_total', 0) or 0))
            for c in range(1, 9):
                _style_data_cell(ws, r, c, c >= 2)

        # 合计行
        total_r = curr_row + len(by_project) + 1
        total_income_sum = sum(p.get('income', 0) or 0 for p in by_project)
        total_expense_sum = sum(p.get('expense', 0) or 0 for p in by_project)
        sum_vals = [
            '合计',
            sum(p.get('task_count', 0) or 0 for p in by_project),
            sum(p.get('task_done', 0) or 0 for p in by_project),
            total_income_sum,
            total_expense_sum,
            total_income_sum - total_expense_sum,
            sum(p.get('acceptance_count', 0) or 0 for p in by_project),
            sum(p.get('acceptance_total', 0) or 0 for p in by_project),
        ]
        bold_font = Font(name='Microsoft YaHei', bold=True, size=10)
        for ci, val in enumerate(sum_vals, 1):
            cell = ws.cell(row=total_r, column=ci, value=val)
            cell.font = bold_font
            cell.alignment = NUM_ALIGN if ci >= 2 else CELL_ALIGN
            cell.border = THIN_BORDER

    _auto_width(ws)
    return _make_excel_response(wb, f'综合报表_{datetime.now().strftime("%Y%m%d")}.xlsx')


# ---------- 单条日志 Excel 导出 ----------

@app.route('/api/logs/<log_id>/excel', methods=['GET'])
def log_export_excel(log_id):
    conn = get_db()
    log = dict_from_row(conn.execute('SELECT * FROM logs WHERE id = ?', (log_id,)).fetchone())
    if not log:
        conn.close()
        return jsonify({'success': False, 'error': '日志不存在'}), 404
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = '施工日志'
    _add_title_sheet(wb, ws, '国标施工日志', f'{log.get("project_name", "") or ""} · 导出时间：{datetime.now().strftime("%Y-%m-%d %H:%M")}')

    # 基本信息
    curr_row = 4
    info_headers = ['日期', '天气', '温度', '风力', '工程名称', '施工单位', '施工部位', '负责人', '记录人']
    for ci, h in enumerate(info_headers, 1):
        ws.cell(row=curr_row, column=ci, value=h)
    _style_header_row(ws, curr_row, len(info_headers))
    info_vals = [
        log.get('date', ''),
        log.get('weather', ''),
        f"{log.get('temp_high', '') or ''}°C ~ {log.get('temp_low', '') or ''}°C",
        log.get('wind', '') or '',
        log.get('project_name', '') or '',
        log.get('unit', '') or '',
        log.get('location', '') or '',
        log.get('manager', '') or '',
        log.get('recorder', '') or '',
    ]
    for ci, val in enumerate(info_vals, 1):
        ws.cell(row=curr_row + 1, column=ci, value=val)
        _style_data_cell(ws, curr_row + 1, ci)

    # 施工记录
    if log.get('production_record'):
        curr_row = curr_row + 3
        ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=9)
        sec = ws.cell(row=curr_row, column=1, value='一、施工记录')
        sec.font = Font(name='Microsoft YaHei', bold=True, size=12)
        curr_row += 1
        ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row + 3, end_column=9)
        content_cell = ws.cell(row=curr_row, column=1, value=log.get('production_record', ''))
        content_cell.font = CELL_FONT
        content_cell.alignment = Alignment(vertical='top', wrap_text=True)
        ws.row_dimensions[curr_row].height = max(80, (log.get('production_record', '') or '').count('\n') * 18 + 30)

    # 技术质量安全
    if log.get('tech_quality_safety'):
        curr_row = curr_row + 5
        ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=9)
        sec2 = ws.cell(row=curr_row, column=1, value='二、技术质量安全')
        sec2.font = Font(name='Microsoft YaHei', bold=True, size=12)
        curr_row += 1
        ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row + 3, end_column=9)
        qs_cell = ws.cell(row=curr_row, column=1, value=log.get('tech_quality_safety', ''))
        qs_cell.font = CELL_FONT
        qs_cell.alignment = Alignment(vertical='top', wrap_text=True)
        ws.row_dimensions[curr_row].height = max(80, (log.get('tech_quality_safety', '') or '').count('\n') * 18 + 30)

    _auto_width(ws, max_width=30)
    # 内容列给足够宽
    ws.column_dimensions['A'].width = 20
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[col_letter].width = 18

    return _make_excel_response(wb, f'施工日志_{log.get("date","") or ""}_{datetime.now().strftime("%Y%m%d")}.xlsx')


# ==================== 返工管理 (reworks) ====================

@app.route('/api/reworks/project/<project_id>', methods=['GET'])
def get_reworks(project_id):
    """获取某工程所有返工记录"""
    conn = get_db()
    items = dicts_from_rows(
        conn.execute('SELECT * FROM reworks WHERE project_id = ? ORDER BY created_at DESC', (project_id,)).fetchall()
    )
    conn.close()
    return jsonify({'success': True, 'data': items})

@app.route('/api/reworks/<rid>', methods=['GET'])
def get_rework(rid):
    """获取单条返工记录"""
    conn = get_db()
    item = dict_from_row(conn.execute('SELECT * FROM reworks WHERE id = ?', (rid,)).fetchone())
    conn.close()
    if not item:
        return jsonify({'success': False, 'error': '返工记录不存在'}), 404
    return jsonify({'success': True, 'data': item})

@app.route('/api/reworks', methods=['POST'])
def create_rework():
    """创建返工记录"""
    data = (request.get_json(silent=True) or {})
    project_id = data.get('project_id', '')
    if not project_id:
        return jsonify({'success': False, 'error': 'project_id 不能为空'}), 400
    rid = data.get('id') or str(int(datetime.now().timestamp() * 1000))
    conn = get_db()
    conn.execute(
        '''INSERT INTO reworks (id, project_id, task_id, name, rework_qty, unit, reason, status, start_date, end_date, remark)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
        (rid, project_id, data.get('task_id', ''), data.get('name', ''),
         data.get('rework_qty', 0), data.get('unit', 'm²'),
         data.get('reason', ''), data.get('status', 'pending'),
         data.get('start_date', ''), data.get('end_date', ''),
         data.get('remark', ''))
    )
    conn.commit()
    item = dict_from_row(conn.execute('SELECT * FROM reworks WHERE id = ?', (rid,)).fetchone())
    conn.close()
    return jsonify({'success': True, 'data': item})

@app.route('/api/reworks/<rid>', methods=['PUT'])
def update_rework(rid):
    """更新返工记录"""
    data = (request.get_json(silent=True) or {})
    conn = get_db()
    conn.execute(
        "UPDATE reworks SET task_id=?, name=?, rework_qty=?, unit=?, reason=?, status=?, start_date=?, end_date=?, remark=?, updated_at=datetime('now','localtime') WHERE id=?",
        (data.get('task_id', ''), data.get('name', ''), data.get('rework_qty', 0),
         data.get('unit', 'm²'), data.get('reason', ''),
         data.get('status', 'pending'), data.get('start_date', ''),
         data.get('end_date', ''), data.get('remark', ''), rid)
    )
    conn.commit()
    item = dict_from_row(conn.execute('SELECT * FROM reworks WHERE id = ?', (rid,)).fetchone())
    conn.close()
    return jsonify({'success': True, 'data': item})

@app.route('/api/reworks/<rid>', methods=['DELETE'])
def delete_rework(rid):
    """删除返工记录"""
    conn = get_db()
    conn.execute('DELETE FROM reworks WHERE id = ?', (rid,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

# ==================== 数据导出/导入 ====================

@app.route('/api/export', methods=['GET'])
def export_data():
    conn = get_db()
    projects = dicts_from_rows(conn.execute('SELECT * FROM projects ORDER BY created_at DESC').fetchall())
    tasks = dicts_from_rows(conn.execute('SELECT * FROM tasks ORDER BY start, id').fetchall())
    logs = dicts_from_rows(conn.execute('SELECT * FROM logs ORDER BY date DESC').fetchall())
    daily_task_logs = dicts_from_rows(conn.execute('SELECT * FROM daily_task_logs ORDER BY log_date').fetchall())
    photos = dicts_from_rows(conn.execute('SELECT * FROM project_photos ORDER BY created_at DESC').fetchall())
    finance = dicts_from_rows(conn.execute('SELECT * FROM misc_finance ORDER BY date DESC').fetchall())
    reworks = dicts_from_rows(conn.execute('SELECT * FROM reworks ORDER BY created_at DESC').fetchall())
    acceptances = dicts_from_rows(conn.execute('SELECT * FROM acceptances ORDER BY date DESC').fetchall())
    materials = dicts_from_rows(conn.execute('SELECT * FROM materials ORDER BY created_at DESC').fetchall())
    equipments = dicts_from_rows(conn.execute('SELECT * FROM equipments ORDER BY created_at DESC').fetchall())
    teams = dicts_from_rows(conn.execute('SELECT * FROM teams ORDER BY created_at DESC').fetchall())
    conn.close()
    return jsonify({
        'success': True,
        'data': {
            'version': '2.4',
            'exported_at': datetime.now().isoformat(),
            'projects': projects,
            'tasks': tasks,
            'logs': logs,
            'daily_task_logs': daily_task_logs,
            'photos': photos,
            'finance': finance,
            'reworks': reworks,
            'acceptances': acceptances,
            'materials': materials,
            'equipments': equipments,
            'teams': teams
        }
    })

@app.route('/api/export/excel', methods=['GET'])
def export_data_excel():
    """全库数据导出为 Excel"""
    conn = get_db()
    projects = dicts_from_rows(conn.execute('SELECT * FROM projects ORDER BY created_at DESC').fetchall())
    tasks = dicts_from_rows(conn.execute('SELECT * FROM tasks ORDER BY start, id').fetchall())
    logs = dicts_from_rows(conn.execute('SELECT * FROM logs ORDER BY date DESC').fetchall())
    daily_task_logs = dicts_from_rows(conn.execute('SELECT * FROM daily_task_logs ORDER BY log_date').fetchall())
    photos = dicts_from_rows(conn.execute('SELECT * FROM project_photos ORDER BY created_at DESC').fetchall())
    finance = dicts_from_rows(conn.execute('SELECT * FROM misc_finance ORDER BY date DESC').fetchall())
    reworks = dicts_from_rows(conn.execute('SELECT * FROM reworks ORDER BY created_at DESC').fetchall())
    acceptances = dicts_from_rows(conn.execute('SELECT * FROM acceptances ORDER BY date DESC').fetchall())
    materials = dicts_from_rows(conn.execute('SELECT * FROM materials ORDER BY created_at DESC').fetchall())
    equipments = dicts_from_rows(conn.execute('SELECT * FROM equipments ORDER BY created_at DESC').fetchall())
    teams = dicts_from_rows(conn.execute('SELECT * FROM teams ORDER BY created_at DESC').fetchall())
    conn.close()

    def _write_sheet(ws, title, rows, columns):
        """通用写入 sheet 函数"""
        _add_title_sheet(wb, ws, title, f'共 {len(rows)} 条记录 · {datetime.now().strftime("%Y-%m-%d %H:%M")}')
        curr = 4
        for ci, h in enumerate(columns, 1):
            ws.cell(row=curr, column=ci, value=h)
        _style_header_row(ws, curr, len(columns))
        for ri, row in enumerate(rows):
            r = curr + 1 + ri
            for ci, key in enumerate(columns, 1):
                val = row.get(key, '')
                if val is None:
                    val = ''
                ws.cell(row=r, column=ci, value=val)
                _style_data_cell(ws, r, ci, isinstance(val, (int, float)))
        _auto_width(ws)

    wb = Workbook()
    # 工程
    ws = wb.active
    ws.title = '工程'
    _write_sheet(ws, '工程列表', projects, ['id', 'name', 'company', 'manager', 'recorder', 'status',
        'start_date', 'end_date', 'address', 'area', 'budget', 'remark'])

    # 任务
    if tasks:
        ws2 = wb.create_sheet('任务')
        _write_sheet(ws2, '任务列表', tasks, ['id', 'project_id', 'name', 'category', 'start', 'end',
            'status', 'team', 'location', 'desc', 'quantity', 'materials', 'equipment', 'quality', 'safety'])

    # 施工日志
    if logs:
        ws3 = wb.create_sheet('施工日志')
        _write_sheet(ws3, '施工日志', logs, ['id', 'project_id', 'date', 'weather', 'temp_high', 'temp_low',
            'wind', 'location', 'unit', 'manager', 'recorder', 'production_record', 'tech_quality_safety',
            'materials', 'equipments', 'daily_task_log_ids', 'incident', 'remark'])

    # 每日任务打卡
    if daily_task_logs:
        ws4 = wb.create_sheet('每日任务打卡')
        _write_sheet(ws4, '每日任务打卡', daily_task_logs, ['id', 'task_id', 'log_date', 'team',
            'worker_count', 'content', 'materials', 'equipments'])

    # 财务
    if finance:
        ws5 = wb.create_sheet('财务')
        _write_sheet(ws5, '财务记录', finance, ['id', 'project_ref', 'date', 'flow_type', 'source',
            'category', 'amount', 'remark', 'status'])

    # 收方
    if acceptances:
        ws6 = wb.create_sheet('收方')
        _write_sheet(ws6, '收方记录', acceptances, ['id', 'project_id', 'name', 'date', 'design_qty',
            'actual_qty', 'unit', 'unit_price', 'total_price', 'status'])

    # 班组
    if teams:
        ws7 = wb.create_sheet('班组')
        _write_sheet(ws7, '班组信息', teams, ['id', 'name', 'worker_count', 'specialty', 'leader', 'phone'])

    # 材料
    if materials:
        ws8 = wb.create_sheet('材料')
        _write_sheet(ws8, '材料库存', materials, ['id', 'project_id', 'name', 'spec', 'quantity', 'unit', 'status'])

    # 设备
    if equipments:
        ws9 = wb.create_sheet('设备')
        _write_sheet(ws9, '设备清单', equipments, ['id', 'project_id', 'name', 'model', 'count', 'status'])

    # 返工
    if reworks:
        ws10 = wb.create_sheet('返工')
        _write_sheet(ws10, '返工记录', reworks, ['id', 'project_id', 'task_id', 'name', 'rework_qty',
            'unit', 'reason', 'status', 'start_date', 'end_date', 'remark'])

    # 影像
    if photos:
        ws11 = wb.create_sheet('影像')
        _write_sheet(ws11, '工程影像', photos, ['id', 'project_id', 'category', 'url', 'note', 'created_at'])

    return _make_excel_response(wb, f'施工日志全库数据_{datetime.now().strftime("%Y%m%d")}.xlsx')

@app.route('/api/import', methods=['POST'])
def import_data():
    data = (request.get_json(silent=True) or {})
    if not data:
        return jsonify({'success': False, 'error': '无效数据'}), 400
    # 兼容前端直接传 data 对象 或 {success:true, data:{...}} 包装
    if 'data' in data and isinstance(data['data'], dict):
        data = data['data']
    conn = get_db()
    # 清空现有数据
    conn.execute('DELETE FROM daily_task_logs')
    conn.execute('DELETE FROM logs')
    conn.execute('DELETE FROM acceptances')
    conn.execute('DELETE FROM project_photos')
    conn.execute('DELETE FROM misc_finance')
    conn.execute('DELETE FROM reworks')
    conn.execute('DELETE FROM materials')
    conn.execute('DELETE FROM equipments')
    conn.execute('DELETE FROM teams')
    conn.execute('DELETE FROM tasks')
    conn.execute('DELETE FROM projects')
    # 导入新数据
    for p in data.get('projects', []):
        conn.execute(
            'INSERT INTO projects (id, name, type, company, client, address, manager, recorder, duration, start_date, end_date, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
            (p.get('id'), p.get('name', ''), p.get('type', 'custom'), p.get('company', ''),
             p.get('client', ''), p.get('address', ''), p.get('manager', ''), p.get('recorder', ''),
             p.get('duration', 0), p.get('start_date', ''), p.get('end_date', ''),
             p.get('created_at', ''), p.get('updated_at', ''))
        )
    for t in data.get('tasks', []):
        conn.execute(
            '''INSERT INTO tasks (id, project_id, category, name, location, start, end, team, workers, status, description, remark, created_at, updated_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
            (t.get('id'), t.get('project_id', ''), t.get('category', ''), t.get('name', ''),
             t.get('location', ''), t.get('start', ''), t.get('end', ''),
             t.get('team', ''), t.get('workers', 0), t.get('status', 'pending'),
             t.get('description', ''), t.get('remark', ''),
             t.get('created_at', ''), t.get('updated_at', ''))
        )
    for l in data.get('logs', []):
        conn.execute(
            '''INSERT INTO logs (id, project_id, project_name, unit, date, weather,
                temp_high, temp_low, wind, location, incident,
                production_record, tech_quality_safety, manager, recorder,
                materials, equipments, daily_task_log_ids, created_at, updated_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
            (l.get('id'), l.get('project_id', ''), l.get('project_name', ''), l.get('unit', ''),
             l.get('date', ''), l.get('weather', ''),
             l.get('temp_high', ''), l.get('temp_low', ''), l.get('wind', ''),
             l.get('location', ''), l.get('incident', ''),
             l.get('production_record', ''), l.get('tech_quality_safety', ''),
             l.get('manager', ''), l.get('recorder', ''),
             l.get('materials', ''), l.get('equipments', ''),
             l.get('daily_task_log_ids', ''),
             l.get('created_at', ''), l.get('updated_at', ''))
        )
    for dtl in data.get('daily_task_logs', []):
        conn.execute(
            '''INSERT INTO daily_task_logs (id, project_id, task_id, log_date, content, weather, team, worker_count, materials, equipments, created_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
            (dtl.get('id'), dtl.get('project_id', ''), dtl.get('task_id', ''),
             dtl.get('log_date', ''), dtl.get('content', ''),
             dtl.get('weather', ''), dtl.get('team', ''),
             dtl.get('worker_count', 0),
             dtl.get('materials', ''), dtl.get('equipments', ''),
             dtl.get('created_at', ''))
        )
    for ph in data.get('photos', []):
        conn.execute(
            '''INSERT INTO project_photos (id, project_id, task_id, category, filename, data, thumbnail, file_size, mime_type, description, taken_at, created_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
            (ph.get('id'), ph.get('project_id', ''), ph.get('task_id', ''), ph.get('category', ''),
             ph.get('filename', ''), ph.get('data', ''), ph.get('thumbnail', ''),
             ph.get('file_size', 0), ph.get('mime_type', 'image/jpeg'),
             ph.get('description', ''), ph.get('taken_at', ''), ph.get('created_at', ''))
        )
    for fn in data.get('finance', []):
        conn.execute(
            '''INSERT INTO misc_finance (id, project_ref, flow_type, source, category, amount, description, date, carryover_mode, carryover_month, carryover_ref_id, status, created_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
            (fn.get('id'), fn.get('project_ref', ''), fn.get('flow_type', 'expense'),
             fn.get('source', 'personal'), fn.get('category', '零星材料'),
             fn.get('amount', 0), fn.get('description', ''), fn.get('date', ''),
             fn.get('carryover_mode', ''), fn.get('carryover_month', ''),
             fn.get('carryover_ref_id', ''), fn.get('status', 'active'),
             fn.get('created_at', ''))
        )
    for rw in data.get('reworks', []):
        conn.execute(
            '''INSERT INTO reworks (id, project_id, task_id, name, rework_qty, unit, reason, status, start_date, end_date, remark, created_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
            (rw.get('id'), rw.get('project_id', ''), rw.get('task_id', ''),
             rw.get('name', ''), rw.get('rework_qty', 0), rw.get('unit', 'm²'),
             rw.get('reason', ''), rw.get('status', 'pending'),
             rw.get('start_date', ''), rw.get('end_date', ''),
             rw.get('remark', ''), rw.get('created_at', ''))
        )
    for a in data.get('acceptances', []):
        conn.execute(
            '''INSERT INTO acceptances (id, project_id, task_id, rework_task_id, acceptance_type,
                name, location, unit, basis, design_qty, actual_qty, unit_price, total_price,
                calc_formula, quantity_type, status, date, remark, created_at, updated_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
            (a.get('id'), a.get('project_id', ''), a.get('task_id', ''),
             a.get('rework_task_id', ''), a.get('acceptance_type', 'normal'),
             a.get('name', ''), a.get('location', ''), a.get('unit', ''),
             a.get('basis', ''), a.get('design_qty', 0), a.get('actual_qty', 0),
             a.get('unit_price', 0), a.get('total_price', 0),
             a.get('calc_formula', ''), a.get('quantity_type', ''),
             a.get('status', 'pending'), a.get('date', ''), a.get('remark', ''),
             a.get('created_at', ''), a.get('updated_at', ''))
        )
    for m in data.get('materials', []):
        conn.execute(
            '''INSERT INTO materials (id, name, spec, unit, quantity, min_quantity, supplier, status, remark, created_at, updated_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
            (m.get('id'), m.get('name', ''),
             m.get('spec', ''), m.get('unit', ''), m.get('quantity', 0),
             m.get('min_quantity', 0), m.get('supplier', ''),
             m.get('status', 'in_stock'), m.get('remark', ''),
             m.get('created_at', ''), m.get('updated_at', ''))
        )
    for eq in data.get('equipments', []):
        conn.execute(
            '''INSERT INTO equipments (id, name, model, count, status, remark, created_at, updated_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
            (eq.get('id'), eq.get('name', ''),
             eq.get('model', ''), eq.get('count', 1), eq.get('status', 'normal'),
             eq.get('remark', ''), eq.get('created_at', ''), eq.get('updated_at', ''))
        )
    for tm in data.get('teams', []):
        conn.execute(
            '''INSERT INTO teams (id, name, leader, phone, specialty, worker_count, remark, created_at, updated_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''',
            (tm.get('id'), tm.get('name', ''),
             tm.get('leader', ''), tm.get('phone', ''), tm.get('specialty', ''),
             tm.get('worker_count', 0), tm.get('remark', ''),
             tm.get('created_at', ''), tm.get('updated_at', ''))
        )
    conn.commit()
    conn.close()
    return jsonify({'success': True})

# ==================== 数据重置 ====================

@app.route('/api/reset', methods=['POST'])
def reset_data():
    conn = get_db()
    conn.execute('DELETE FROM daily_task_logs')
    conn.execute('DELETE FROM logs')
    conn.execute('DELETE FROM acceptances')
    conn.execute('DELETE FROM project_photos')
    conn.execute('DELETE FROM misc_finance')
    conn.execute('DELETE FROM reworks')
    conn.execute('DELETE FROM materials')
    conn.execute('DELETE FROM equipments')
    conn.execute('DELETE FROM teams')
    conn.execute('DELETE FROM tasks')
    conn.execute('DELETE FROM projects')
    conn.commit()
    conn.close()
    return jsonify({'success': True})

# ==================== 排工考勤系统（独立数据） ====================
PAIBAN_DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'paiban', 'paiban.db')

def init_paiban_db():
    """初始化排工独立数据库"""
    os.makedirs(os.path.dirname(PAIBAN_DB_PATH), exist_ok=True)
    conn = sqlite3.connect(PAIBAN_DB_PATH)
    conn.execute('''CREATE TABLE IF NOT EXISTS paiban_data (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        workers TEXT DEFAULT '[]',
        tasks TEXT DEFAULT '[]',
        plan_data TEXT DEFAULT '{}',
        attend_data TEXT DEFAULT '{}',
        docs TEXT DEFAULT '[]',
        archives TEXT DEFAULT '{}',
        updated_at TEXT DEFAULT (datetime('now','localtime'))
    )''')
    # 确保只有一行数据（单用户模式）
    existing = conn.execute('SELECT id FROM paiban_data LIMIT 1').fetchone()
    if not existing:
        conn.execute('INSERT INTO paiban_data (workers, tasks, plan_data, attend_data, docs, archives) VALUES (?, ?, ?, ?, ?, ?)',
                     ('[]', '[]', '{}', '{}', '[]', '{}'))
    conn.commit()
    conn.close()

@app.route('/api/paiban/data', methods=['GET'])
def paiban_get_data():
    """获取排工全部数据"""
    conn = sqlite3.connect(PAIBAN_DB_PATH)
    row = conn.execute('SELECT workers, tasks, plan_data, attend_data, docs, archives, updated_at FROM paiban_data LIMIT 1').fetchone()
    conn.close()
    if not row:
        return jsonify({'success': False, 'error': '无数据'}), 404
    return jsonify({'success': True, 'data': {
        'workers': json.loads(row[0] or '[]'),
        'tasks': json.loads(row[1] or '[]'),
        'planData': json.loads(row[2] or '{}'),
        'attendData': json.loads(row[3] or '{}'),
        'docs': json.loads(row[4] or '[]'),
        'archives': json.loads(row[5] or '{}'),
        'updated_at': row[6]
    }})

@app.route('/api/paiban/data', methods=['POST'])
def paiban_save_data():
    """保存排工全部数据"""
    data = request.get_json(silent=True) or {}
    conn = sqlite3.connect(PAIBAN_DB_PATH)
    conn.execute('''UPDATE paiban_data SET 
        workers = ?, tasks = ?, plan_data = ?, attend_data = ?, docs = ?, archives = ?,
        updated_at = datetime('now','localtime')
    ''', (
        json.dumps(data.get('workers', []), ensure_ascii=False),
        json.dumps(data.get('tasks', []), ensure_ascii=False),
        json.dumps(data.get('planData', {}), ensure_ascii=False),
        json.dumps(data.get('attendData', {}), ensure_ascii=False),
        json.dumps(data.get('docs', []), ensure_ascii=False),
        json.dumps(data.get('archives', {}), ensure_ascii=False)
    ))
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')})

# ==================== Git自动拉取 ====================

@app.route('/api/git-pull', methods=['POST'])
def git_pull():
    """自动拉取最新代码（供定时任务调用）"""
    reload_msg = ''
    try:
        project_dir = os.path.dirname(os.path.abspath(__file__))
        result = subprocess.run(
            ['git', 'pull'],
            cwd=project_dir,
            capture_output=True,
            text=True,
            timeout=30
        )
        # 代码有更新时，自动触发 Reload（修改 WSGI 文件时间戳）
        if result.returncode == 0 and 'Already up to date' not in result.stdout:
            wsgi_path = '/var/www/slhfwq_pythonanywhere_com_wsgi.py'
            if os.path.exists(wsgi_path):
                os.utime(wsgi_path, None)
                reload_msg = '，已自动触发Reload'
        return jsonify({
            'success': result.returncode == 0,
            'stdout': result.stdout.strip(),
            'stderr': result.stderr.strip(),
            'returncode': result.returncode,
            'reload': 'Reload已触发' if reload_msg else '无需Reload'
        })
    except subprocess.TimeoutExpired:
        return jsonify({'success': False, 'error': 'Git pull 超时（30秒）'}), 408
    except FileNotFoundError:
        return jsonify({'success': False, 'error': '未找到 git 命令，请确认 git 已安装'}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': f'执行失败: {str(e)}'}), 500

# ==================== 排工系统更新接口 ====================

PAIBAN_VERSION = {
    "versionCode": 1,
    "versionName": "1.0",
    "downloadUrl": "/paiban/app-debug.apk"
}

@app.route('/paiban/api/version')
def paiban_version():
    """排工APK版本检测接口"""
    return jsonify(PAIBAN_VERSION)

@app.route('/paiban/app-debug.apk')
def paiban_download_apk():
    """排工APK下载"""
    apk_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'paiban', 'app-debug.apk')
    if not os.path.exists(apk_path):
        return jsonify({'error': 'APK文件不存在'}), 404
    return send_from_directory('paiban', 'app-debug.apk', as_attachment=True, download_name='枫叶管理.apk')

# ==================== GPA系统（独立子系统，使用独立数据库） ====================

GPA_DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'gpa', 'gpa.db')

DEFAULT_SEMESTERS = [
    {'id': 'sem-1', 'name': '大一上学期', 'startDate': '', 'endDate': '', 'note': ''},
    {'id': 'sem-2', 'name': '大一下学期', 'startDate': '', 'endDate': '', 'note': ''},
    {'id': 'sem-3', 'name': '大二上学期', 'startDate': '', 'endDate': '', 'note': ''},
    {'id': 'sem-4', 'name': '大二下学期', 'startDate': '', 'endDate': '', 'note': ''},
    {'id': 'sem-5', 'name': '大三上学期', 'startDate': '', 'endDate': '', 'note': ''},
    {'id': 'sem-6', 'name': '大三下学期', 'startDate': '', 'endDate': '', 'note': ''},
    {'id': 'sem-7', 'name': '大四上学期', 'startDate': '', 'endDate': '', 'note': ''},
    {'id': 'sem-8', 'name': '大四下学期', 'startDate': '', 'endDate': '', 'note': ''}
]

DEFAULT_COURSES_SEM1 = [
    {'id': 'c-1', 'semesterId': 'sem-1', 'name': '大学生心理健康', 'credit': 2, 'score': 91, 'category': '公共必修'},
    {'id': 'c-2', 'semesterId': 'sem-1', 'name': '党史', 'credit': 1, 'score': 89, 'category': '公共必修'},
    {'id': 'c-3', 'semesterId': 'sem-1', 'name': '高等数学C1', 'credit': 4, 'score': 85, 'category': '专业必修'},
    {'id': 'c-4', 'semesterId': 'sem-1', 'name': '高级交际英语1', 'credit': 3, 'score': 75, 'category': '公共必修'},
    {'id': 'c-5', 'semesterId': 'sem-1', 'name': '国家安全教育', 'credit': 1, 'score': 80, 'category': '公共必修'},
    {'id': 'c-6', 'semesterId': 'sem-1', 'name': '逻辑学导论', 'credit': 3, 'score': 60, 'category': '专业必修'},
    {'id': 'c-7', 'semesterId': 'sem-1', 'name': '思想道德与法治', 'credit': 2.5, 'score': 90, 'category': '公共必修'},
    {'id': 'c-8', 'semesterId': 'sem-1', 'name': '素质体育1', 'credit': 1, 'score': 90, 'category': '公共必修'},
    {'id': 'c-9', 'semesterId': 'sem-1', 'name': '微观经济学', 'credit': 3, 'score': 80, 'category': '专业必修'},
    {'id': 'c-10', 'semesterId': 'sem-1', 'name': '形势与政策1', 'credit': 0.5, 'score': 96, 'category': '公共必修'},
    {'id': 'c-11', 'semesterId': 'sem-1', 'name': '政治学原理', 'credit': 3, 'score': 84, 'category': '专业必修'}
]

def init_gpa_db():
    """初始化GPA独立数据库，自动迁移预置数据"""
    os.makedirs(os.path.dirname(GPA_DB_PATH), exist_ok=True)
    conn = sqlite3.connect(GPA_DB_PATH)
    conn.execute('''CREATE TABLE IF NOT EXISTS gpa_data (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        semesters TEXT DEFAULT '[]',
        courses TEXT DEFAULT '[]',
        updated_at TEXT DEFAULT (datetime('now','localtime'))
    )''')
    existing = conn.execute('SELECT id, semesters, courses FROM gpa_data LIMIT 1').fetchone()
    if not existing:
        conn.execute('INSERT INTO gpa_data (semesters, courses) VALUES (?, ?)',
                     (json.dumps(DEFAULT_SEMESTERS, ensure_ascii=False),
                      json.dumps(DEFAULT_COURSES_SEM1, ensure_ascii=False)))
    else:
        # 合并预置学期（仅添加缺失的）
        cur_semesters = json.loads(existing[1] or '[]')
        sem_ids = {s['id'] for s in cur_semesters}
        changed = False
        for sem in DEFAULT_SEMESTERS:
            if sem['id'] not in sem_ids:
                cur_semesters.append(sem)
                sem_ids.add(sem['id'])
                changed = True
        # 合并大一上预置课程（仅在完全没有sem-1课程时添加）
        cur_courses = json.loads(existing[2] or '[]')
        has_sem1 = any(c.get('semesterId') == 'sem-1' for c in cur_courses)
        if not has_sem1:
            cur_courses.extend(DEFAULT_COURSES_SEM1)
            changed = True
        if changed:
            conn.execute('''UPDATE gpa_data SET
                semesters = ?, courses = ?,
                updated_at = datetime('now','localtime')
            ''', (json.dumps(cur_semesters, ensure_ascii=False),
                  json.dumps(cur_courses, ensure_ascii=False)))
    conn.commit()
    conn.close()

@app.route('/api/gpa/data', methods=['GET'])
def gpa_get_data():
    """获取GPA全部数据（懒迁移：无sem-1课程时自动补）"""
    conn = sqlite3.connect(GPA_DB_PATH)
    row = conn.execute('SELECT id, semesters, courses, updated_at FROM gpa_data LIMIT 1').fetchone()
    if not row:
        conn.close()
        return jsonify({'success': False, 'error': '无数据'}), 404
    cur_courses = json.loads(row[2] or '[]')
    cur_semesters = json.loads(row[1] or '[]')
    # 懒迁移：缺乏sem-1课程时自动补充
    has_sem1 = any(c.get('semesterId') == 'sem-1' for c in cur_courses)
    any_sem = any(c.get('semesterId') for c in cur_courses)
    if not has_sem1 and not any_sem and any(s.get('id') == 'sem-1' for s in cur_semesters):
        cur_courses.extend(DEFAULT_COURSES_SEM1)
        conn.execute('''UPDATE gpa_data SET courses = ?, updated_at = datetime('now','localtime')''',
                     (json.dumps(cur_courses, ensure_ascii=False),))
        conn.commit()
    conn.close()
    return jsonify({'success': True, 'data': {
        'semesters': cur_semesters,
        'courses': cur_courses,
        'updated_at': row[3]
    }})

@app.route('/api/gpa/data', methods=['POST'])
def gpa_save_data():
    """保存GPA全部数据"""
    data = request.get_json(silent=True) or {}
    conn = sqlite3.connect(GPA_DB_PATH)
    conn.execute('''UPDATE gpa_data SET
        semesters = ?, courses = ?,
        updated_at = datetime('now','localtime')
    ''', (
        json.dumps(data.get('semesters', []), ensure_ascii=False),
        json.dumps(data.get('courses', []), ensure_ascii=False),
    ))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

# ==================== 前端路由 ====================

@app.route('/gpa')
def gpa_index():
    """大学GPA记录系统（独立子系统）"""
    return send_from_directory('gpa', 'index.html')

@app.route('/paiban')
def paiban_index():
    """排工考勤系统（独立子系统，数据不交叉）"""
    return send_from_directory('paiban', 'index.html')

@app.route('/')
def index():
    return send_from_directory('public', 'index.html')

@app.route('/<path:path>')
def static_files(path):
    return send_from_directory('public', path)

# ==================== 启动 ====================

# 应用导入时自动初始化数据库
init_db()
init_paiban_db()
init_gpa_db()

if __name__ == '__main__':
    print('施工日志系统: http://localhost:5000')
    print('排工考勤系统: http://localhost:5000/paiban')
    app.run(host='0.0.0.0', port=5000, debug=True)
